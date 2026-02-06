import streamlit as st
import pandas as pd
import time
from datetime import datetime
import numpy as np

# --- DATABASE VE CORE IMPORTLARI ---
from app.core.database import fetch_data, add_data, get_conn, update_data
from app.core.config import INPUT_LIMITS, TERMS, get_limit
from app.core.error_handling import error_handler, log_info, log_warning, ERROR_HANDLING_AVAILABLE
from app.core.components import render_help_button
from app.core.languages import t

# Rapor modÃ¼lÃ¼ (Hata Ã¶nleyici)
try:
    from app.modules.reports import download_styled_excel as shared_download
except ImportError:
    def shared_download(*args): pass
        

# --------------------------------------------------------------------------
# YARDIMCI FONKSÄ°YONLAR (Dashboard BaÄŸÄ±mlÄ±lÄ±ÄŸÄ±nÄ± KaldÄ±rmak Ä°Ã§in Buraya Eklendi)
# --------------------------------------------------------------------------
def update_tavli_record_backend(original_record, new_data):
    """
    TavlÄ± analiz kaydÄ±nÄ± gÃ¼nceller ve silolar tablosundaki stoklarÄ± senkronize eder.
    """
    try:
        conn = get_conn()
        df_tavli = fetch_data("tavli_analiz")
        
        # KaydÄ± bul (Tarih ve Silo ismine gÃ¶re eÅŸleÅŸtirme - ID olmadÄ±ÄŸÄ± iÃ§in)
        # Not: GerÃ§ek sistemde ID olmasÄ± daha iyidir ama mevcut yapÄ±da timestamp kullanÄ±yoruz.
        match_idx = df_tavli[
            (df_tavli['tarih'].astype(str) == str(original_record['tarih'])) & 
            (df_tavli['silo_isim'] == original_record['silo_isim'])
        ].index
        
        if len(match_idx) == 0:
            return False, "KayÄ±t veritabanÄ±nda bulunamadÄ±."
            
        idx = match_idx[0]
        
        # --- STOK DÃœZELTME MANTIÄI ---
        # EÄŸer Silo veya Tonaj deÄŸiÅŸtiyse, eski stoÄŸu geri al, yenisini iÅŸle.
        old_silo = original_record['silo_isim']
        new_silo = new_data['silo_isim']
        old_tonaj = float(original_record['analiz_tonaj'])
        new_tonaj = float(new_data['analiz_tonaj'])
        
        if old_silo != new_silo or old_tonaj != new_tonaj:
            # 1. Eski silodan dÃ¼ÅŸ (Reverse operation)
            # update_tavli_bugday_stok fonksiyonunu 'cikar' modunda eski veriyle Ã§alÄ±ÅŸtÄ±r
            update_tavli_bugday_stok(old_silo, old_tonaj, "cikar")
            
            # 2. Yeni siloya ekle
            update_tavli_bugday_stok(new_silo, new_tonaj, "ekle")
            
        # --- VERÄ° GÃœNCELLEME ---
        for key, val in new_data.items():
            df_tavli.at[idx, key] = val
            
        conn.update(worksheet="tavli_analiz", data=df_tavli)
        return True, "âœ… TavlÄ± analiz ve stok kartlarÄ± baÅŸarÄ±yla gÃ¼ncellendi."
        
    except Exception as e:
        return False, f"GÃ¼ncelleme HatasÄ±: {str(e)}"

def delete_tavli_record_backend(record):
    """
    TavlÄ± analiz kaydÄ±nÄ± siler ve stoÄŸu dÃ¼ÅŸer.
    """
    try:
        conn = get_conn()
        df_tavli = fetch_data("tavli_analiz")
        
        # KaydÄ± bul
        mask = (df_tavli['tarih'].astype(str) == str(record['tarih'])) & \
               (df_tavli['silo_isim'] == record['silo_isim'])
               
        if not mask.any():
            return False, "Silinecek kayÄ±t bulunamadÄ±."
            
        # 1. Stoktan DÃ¼ÅŸ (Bu analiz silindiÄŸi iÃ§in, o tavlÄ± miktar da yok sayÄ±lmalÄ± veya serbest bÄ±rakÄ±lmalÄ±)
        # Not: TavlÄ± stoktan dÃ¼ÅŸÃ¼yoruz Ã§Ã¼nkÃ¼ bu analiz o stoÄŸu "tavlÄ±" olarak iÅŸaretlemiÅŸti.
        update_tavli_bugday_stok(record['silo_isim'], record['analiz_tonaj'], "cikar")
        
        # 2. KaydÄ± Sil
        df_new = df_tavli[~mask]
        conn.update(worksheet="tavli_analiz", data=df_new)
        
        return True, "ğŸ—‘ï¸ KayÄ±t silindi ve stok gÃ¼ncellendi."
    except Exception as e:
        return False, f"Silme HatasÄ±: {str(e)}"
def delete_intake_record(lot_no):
    """
    Bir mal kabul kaydÄ±nÄ± SÄ°LER.
    Profesyonel YaklaÅŸÄ±m: Hem arÅŸivden siler, hem stok hareketini siler, hem de siloyu gÃ¼nceller.
    """
    try:
        conn = get_conn()
        
        # 1. ArÅŸivden Sil
        df_arsiv = fetch_data("bugday_giris_arsivi")
        if not df_arsiv.empty and 'lot_no' in df_arsiv.columns:
            df_arsiv = df_arsiv[df_arsiv['lot_no'] != lot_no]
            update_data("bugday_giris_arsivi", df_arsiv) # update_data yoksa conn.update kullan
        
        # 2. Hareketlerden Sil (Stok DÃ¼ÅŸmesi Ä°Ã§in)
        df_hareket = fetch_data("hareketler")
        if not df_hareket.empty and 'lot_no' in df_hareket.columns:
            df_hareket = df_hareket[df_hareket['lot_no'] != lot_no]
            update_data("hareketler", df_hareket)
            
        # 3. SilolarÄ± Yeniden Hesapla (En Kritik AdÄ±m)
        recalculate_silos_from_logs()
        
        return True, "KayÄ±t ve ilgili stok hareketleri baÅŸarÄ±yla silindi."
    except Exception as e:
        return False, f"Silme hatasÄ±: {str(e)}"

def update_intake_record(old_lot_no, new_data):
    """
    Bir mal kabul kaydÄ±nÄ± GÃœNCELLER (Full Yetkili).
    Silo ismi, tonaj veya analiz deÄŸiÅŸirse stok hareketlerini ve ortalamalarÄ± da dÃ¼zeltir.
    """
    try:
        conn = get_conn()
        
        # 1. ArÅŸivi GÃ¼ncelle (TÃ¼m Detaylar Buraya YazÄ±lÄ±r)
        df_arsiv = fetch_data("bugday_giris_arsivi")
        if not df_arsiv.empty and 'lot_no' in df_arsiv.columns:
            idx_list = df_arsiv.index[df_arsiv['lot_no'] == old_lot_no].tolist()
            if idx_list:
                idx = idx_list[0]
                # Yeni verileri iÅŸle
                for key, val in new_data.items():
                    df_arsiv.at[idx, key] = val
                
                update_data("bugday_giris_arsivi", df_arsiv)
        
        # 2. Hareket Tablosunu GÃ¼ncelle (Senkronizasyon)
        # BurasÄ± Stok HesabÄ± ve PaÃ§al Kalitesi Ä°Ã§in Kritiktir
        df_hareket = fetch_data("hareketler")
        if not df_hareket.empty and 'lot_no' in df_hareket.columns:
            idx_list_h = df_hareket.index[df_hareket['lot_no'] == old_lot_no].tolist()
            if idx_list_h:
                idx_h = idx_list_h[0]
                
                # Hareket tablosundaki karÅŸÄ±lÄ±klarÄ± eÅŸle
                mapping = {
                    'tonaj': 'miktar',          # ArÅŸivdeki 'tonaj' -> Hareketteki 'miktar'
                    'fiyat': 'maliyet',         # ArÅŸivdeki 'fiyat' -> Hareketteki 'maliyet'
                    'silo_isim': 'silo_isim',   # KRÄ°TÄ°K: Silo deÄŸiÅŸirse burasÄ± gÃ¼ncellenir
                    'protein': 'protein',
                    'gluten': 'gluten',
                    'rutubet': 'rutubet',
                    'hektolitre': 'hektolitre',
                    'sedim': 'sedim',
                    'tedarikci': 'tedarikci',
                    'notlar': 'notlar'
                }
                
                for key_arsiv, key_hareket in mapping.items():
                    if key_arsiv in new_data:
                         df_hareket.at[idx_h, key_hareket] = new_data[key_arsiv]
                
                update_data("hareketler", df_hareket)

        # 3. SilolarÄ± Yeniden Hesapla
        # Silo ismi veya tonaj deÄŸiÅŸtiÄŸi iÃ§in tÃ¼m deponun yeniden matematiksel olarak kurgulanmasÄ± gerekir.
        recalculate_silos_from_logs()
        
        return True, "âœ… KayÄ±t baÅŸarÄ±yla gÃ¼ncellendi, stoklar ve ortalamalar eÅŸitlendi."
    except Exception as e:
        return False, f"GÃ¼ncelleme hatasÄ±: {str(e)}"

def draw_silo(fill_ratio, name):
    """Silo gÃ¶rseli Ã§iz"""
    try:
        fill_ratio = float(fill_ratio)
        fill_ratio = max(0.0, min(1.0, fill_ratio))
    except: fill_ratio = 0.0
    
    height = 100
    fill_height = int(height * fill_ratio)
    empty_height = height - fill_height
    
    try:
        if fill_ratio < 0.2: fill_color = "#EF4444"
        elif fill_ratio < 0.5: fill_color = "#3B82F6"
        elif fill_ratio < 0.8: fill_color = "#10B981"
        else: fill_color = "#F59E0B"
    except: fill_color = "#CBD5E1"
    
    svg = f'''<svg width="60" height="{height + 10}">
        <rect x="10" y="5" width="40" height="{height}" rx="5" ry="5" 
              style="fill: #f0f2f6; stroke: #333; stroke-width:2;"/>
        <rect x="10" y="{5 + empty_height}" width="40" height="{fill_height}" 
              rx="5" ry="5" style="fill: {fill_color}; stroke: none;"/>
        <text x="30" y="{height + 5}" font-size="8" text-anchor="middle" 
              fill="#333">{name}</text>
    </svg>'''
    return svg

def get_silo_data():
    """Silo verilerini getir"""
    try:
        df = fetch_data("silolar")
        if df.empty:
            return pd.DataFrame(columns=['isim', 'kapasite', 'mevcut_miktar', 'bugday_cinsi', 'maliyet'])
        # NaN temizliÄŸi
        cols = ['protein', 'gluten', 'rutubet', 'hektolitre', 'sedim', 'maliyet', 'mevcut_miktar', 'kapasite']
        for col in cols:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)
        
        if 'isim' in df.columns:
            df = df.sort_values('isim')
        return df
    except Exception as e:
        st.error(f"Silo verisi hatasÄ±: {e}")
        return pd.DataFrame()

# --------------------------------------------------------------------------
# VERÄ° Ä°ÅLEME FONKSÄ°YONLARI (ORÄ°JÄ°NAL MANTIK - GOOGLE SHEETS ADAPTASYONU)
# --------------------------------------------------------------------------

@error_handler(context="Stok Hareketi Loglama")
def log_stok_hareketi(silo_isim, hareket_tipi, miktar, **kwargs):
    """Stok hareketini logla (TÃœM PARAMETRELER DAHÄ°L)"""
    try:
        unique_id = int(datetime.now().timestamp() * 1000)
        
        # Orijinal koddaki tÃ¼m opsiyonel alanlarÄ± kapsayan yapÄ±
        data = {
            'id': unique_id,
            'silo_isim': silo_isim,
            'hareket_tipi': hareket_tipi,
            'miktar': abs(float(miktar)),
            'tarih': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            # Analiz DeÄŸerleri
            'protein': kwargs.get('protein', 0),
            'gluten': kwargs.get('gluten', 0),
            'rutubet': kwargs.get('rutubet', 0),
            'hektolitre': kwargs.get('hektolitre', 0),
            'sedim': kwargs.get('sedim', 0),
            'maliyet': kwargs.get('maliyet', 0),
            # Lojistik Bilgiler
            'lot_no': kwargs.get('lot_no', ''),
            'tedarikci': kwargs.get('tedarikci', ''),
            'yore': kwargs.get('yore', ''),
            'notlar': kwargs.get('notlar', '')
        }
        return add_data("hareketler", data)
    except Exception as e:
        st.error(f"âŒ Hareket kaydÄ± hatasÄ±: {str(e)}")
        return False

def update_tavli_bugday_stok(silo_isim, eklenen_tonaj, islem_tipi="ekle"):
    """TavlÄ± buÄŸday stokunu gÃ¼ncelle"""
    try:
        conn = get_conn()
        df = fetch_data("silolar")
        if df.empty: return False

        mask = df['isim'] == silo_isim
        if not mask.any(): return False
            
        current = float(df.loc[mask, 'tavli_bugday_stok'].iloc[0]) if pd.notnull(df.loc[mask, 'tavli_bugday_stok'].iloc[0]) else 0.0
        
        if islem_tipi == "ekle":
            yeni_tavli = current + float(eklenen_tonaj)
        elif islem_tipi == "cikar":
            yeni_tavli = max(0, current - float(eklenen_tonaj))
        else: return False
            
        df.loc[mask, 'tavli_bugday_stok'] = yeni_tavli
        conn.update(worksheet="silolar", data=df)
        return True
    except Exception as e:
        st.error(f"TavlÄ± stok gÃ¼ncelleme hatasÄ±: {str(e)}")
        return False

def recalculate_silos_from_logs():
    """
    GeÃ§miÅŸ hareketleri tarayÄ±p silolarÄ± senkronize eder (SQL MantÄ±ÄŸÄ± -> Pandas MantÄ±ÄŸÄ±)
    
    Ã–NEMLÄ°: Bu fonksiyon her mal kabul/Ã§Ä±kÄ±ÅŸtan sonra otomatik Ã§aÄŸrÄ±lÄ±r!
    """
    try:
        # ===== VERÄ°LERÄ° Ã‡EK (FORCE REFRESH) =====
        from app.core.database import update_data, clear_cache
        
        # Cache'i temizle ve taze veri al
        clear_cache("silolar")
        clear_cache("hareketler")
        
        df_silolar = fetch_data("silolar", force_refresh=True)
        df_hareketler = fetch_data("hareketler", force_refresh=True)
        
        if df_silolar.empty:
            st.warning("âš ï¸ Silolar tablosu boÅŸ!")
            return False
        
        # Hareket yoksa silolarÄ± sÄ±fÄ±rla ve Ã§Ä±k
        if df_hareketler.empty:
            st.info("â„¹ï¸ HenÃ¼z hareket kaydÄ± yok, silolar sÄ±fÄ±rlanÄ±yor.")
            df_silolar['mevcut_miktar'] = 0.0
            df_silolar['protein'] = 0.0
            df_silolar['maliyet'] = 0.0
            return update_data("silolar", df_silolar)
        
        # ===== NUMERIC KOLONLARI DÃœZELT =====
        numeric_cols = ['miktar', 'protein', 'maliyet', 'gluten', 'rutubet', 'hektolitre', 'sedim']
        for col in numeric_cols:
            if col in df_hareketler.columns:
                df_hareketler[col] = pd.to_numeric(df_hareketler[col], errors='coerce').fillna(0)
        
        # ===== HER SÄ°LO Ä°Ã‡Ä°N HESAPLA =====
        for index, row in df_silolar.iterrows():
            silo_isim = row['isim']
            
            # Bu silonun hareketlerini filtrele
            silo_moves = df_hareketler[df_hareketler['silo_isim'] == silo_isim].copy()
            
            if silo_moves.empty:
                # Hareket yoksa stoÄŸu sÄ±fÄ±rla
                df_silolar.at[index, 'mevcut_miktar'] = 0.0
                df_silolar.at[index, 'protein'] = 0.0
                df_silolar.at[index, 'maliyet'] = 0.0
                continue
            
            # ===== GÄ°RÄ°Å VE Ã‡IKIÅ AYIR =====
            girisler = silo_moves[silo_moves['hareket_tipi'] == 'GiriÅŸ'].copy()
            cikislar = silo_moves[silo_moves['hareket_tipi'] == 'Ã‡Ä±kÄ±ÅŸ'].copy()
            
            # ===== TOPLAM HESAPLA =====
            toplam_giris = girisler['miktar'].sum() if not girisler.empty else 0.0
            toplam_cikis = cikislar['miktar'].sum() if not cikislar.empty else 0.0
            
            mevcut_miktar = max(0, toplam_giris - toplam_cikis)
            
            # ===== AÄIRLIKLI ORTALAMA (Sadece GiriÅŸlerden) =====
            if not girisler.empty and toplam_giris > 0:
                try:
                    # Protein ortalamasÄ±
                    avg_protein = (girisler['miktar'] * girisler['protein']).sum() / toplam_giris
                    
                    # Maliyet ortalamasÄ±
                    avg_maliyet = (girisler['miktar'] * girisler['maliyet']).sum() / toplam_giris
                    
                    # DiÄŸer parametreler (opsiyonel)
                    avg_gluten = (girisler['miktar'] * girisler['gluten']).sum() / toplam_giris if 'gluten' in girisler.columns else 0
                    avg_rutubet = (girisler['miktar'] * girisler['rutubet']).sum() / toplam_giris if 'rutubet' in girisler.columns else 0
                    avg_hektolitre = (girisler['miktar'] * girisler['hektolitre']).sum() / toplam_giris if 'hektolitre' in girisler.columns else 0
                    avg_sedim = (girisler['miktar'] * girisler['sedim']).sum() / toplam_giris if 'sedim' in girisler.columns else 0
                    
                    # Silo tablosunu gÃ¼ncelle
                    df_silolar.at[index, 'protein'] = avg_protein
                    df_silolar.at[index, 'maliyet'] = avg_maliyet
                    
                    # EÄŸer sÃ¼tunlar varsa diÄŸerlerini de gÃ¼ncelle
                    if 'gluten' in df_silolar.columns:
                        df_silolar.at[index, 'gluten'] = avg_gluten
                    if 'rutubet' in df_silolar.columns:
                        df_silolar.at[index, 'rutubet'] = avg_rutubet
                    if 'hektolitre' in df_silolar.columns:
                        df_silolar.at[index, 'hektolitre'] = avg_hektolitre
                    if 'sedim' in df_silolar.columns:
                        df_silolar.at[index, 'sedim'] = avg_sedim
                    
                except Exception as calc_err:
                    st.warning(f"âš ï¸ {silo_isim} iÃ§in ortalama hesaplanamadÄ±: {calc_err}")
                    # Hesaplama hatasÄ± olsa bile miktar gÃ¼ncellensin
                    pass
            else:
                # GiriÅŸ yoksa varsayÄ±lan deÄŸerler
                df_silolar.at[index, 'protein'] = 0.0
                df_silolar.at[index, 'maliyet'] = 0.0
            
            # ===== MEVCUT MÄ°KTARI GÃœNCELLE =====
            df_silolar.at[index, 'mevcut_miktar'] = mevcut_miktar
        
        # ===== GOOGLE SHEETS'E KAYDET (YENÄ° METODUMUZLA) =====
        if update_data("silolar", df_silolar):
            # BaÅŸarÄ± mesajÄ± (opsiyonel - Ã§ok fazla gÃ¶sterilirse yorucu olur)
            # st.success("âœ… Silo stoklarÄ± gÃ¼ncellendi!")
            return True
        else:
            st.error("âŒ Silo gÃ¼ncellemesi baÅŸarÄ±sÄ±z!")
            return False
        
    except Exception as e:
        st.error(f"âŒ Silo hesaplama hatasÄ±: {str(e)}")
        import traceback
        st.code(traceback.format_exc())
        return False

def add_to_bugday_giris_arsivi(lot_no, **kwargs):
    """BuÄŸday giriÅŸini arÅŸive ekle (DETAYLI KAYIT)"""
    try:
        # kwargs iÃ§inde orijinal kodundaki tÃ¼m parametreler gelecek:
        # tarih, bugday_cinsi, tedarikci, yore, plaka, tonaj, fiyat, silo_isim,
        # hektolitre, protein, rutubet, gluten, gluten_index, sedim, gecikmeli_sedim,
        # sune, kirik_ciliz, yabanci_tane, notlar
        
        data = {'lot_no': lot_no, **kwargs}
        return add_data("bugday_giris_arsivi", data)
    except Exception as e:
        st.error(f"âŒ ArÅŸiv hatasÄ±: {str(e)}")
        return False

def get_movements():
    """Stok hareketlerini detaylÄ± getir (ArÅŸiv ile JOIN iÅŸlemi)"""
    try:
        df_h = fetch_data("hareketler")
        df_a = fetch_data("bugday_giris_arsivi")
        
        # BOÅLUK KONTROLÃœ
        if df_h.empty:
            st.warning("ğŸ” Hareketler tablosu boÅŸ!")
            return pd.DataFrame()
        
        # DEBUG: SÃ¼tunlarÄ± gÃ¶ster (geÃ§ici - sonra silebilirsin)
        # st.info(f"Hareketler sÃ¼tunlarÄ±: {list(df_h.columns)}")
        # if not df_a.empty:
        #     st.info(f"ArÅŸiv sÃ¼tunlarÄ±: {list(df_a.columns)}")
        
        # ArÅŸiv yoksa hareketleri olduÄŸu gibi dÃ¶ndÃ¼r
        if df_a.empty:
            if 'tarih' in df_h.columns:
                df_h['tarih'] = pd.to_datetime(df_h['tarih'], errors='coerce')
                df_h = df_h.sort_values('tarih', ascending=False)
            return df_h
        
        # ===== LOT_NO KONTROLÃœ =====
        if 'lot_no' not in df_h.columns:
            st.error("âŒ 'lot_no' sÃ¼tunu hareketler tablosunda bulunamadÄ±!")
            # lot_no yoksa hareketleri olduÄŸu gibi gÃ¶ster
            return df_h
        
        if 'lot_no' not in df_a.columns:
            st.warning("âš ï¸ 'lot_no' sÃ¼tunu arÅŸiv tablosunda bulunamadÄ±!")
            return df_h
        
        # ===== ARÅÄ°VDEN ALINACAK SÃœTUNLARI BELÄ°RLE (Mevcut olanlarÄ± al) =====
        arsiv_kolonlar = ['lot_no']  # lot_no kesin olmalÄ±
        
        # Ä°steÄŸe baÄŸlÄ± sÃ¼tunlarÄ± ekle (varsa)
        optional_cols = [
            'tedarikci', 'yore', 'plaka', 'bugday_cinsi', 
            'sune', 'kirik_ciliz', 'yabanci_tane', 
            'gluten_index', 'gecikmeli_sedim'
        ]
        
        for col in optional_cols:
            if col in df_a.columns:
                arsiv_kolonlar.append(col)
        
        # ===== PANDAS MERGE (LEFT JOIN) =====
        merged = pd.merge(
            df_h, 
            df_a[arsiv_kolonlar], 
            on='lot_no', 
            how='left',  # Sol tablodaki (hareketler) tÃ¼m kayÄ±tlarÄ± koru
            suffixes=('', '_arsiv')
        )
        
        # ===== Ã‡AKIÅAN SÃœTUNLARI BÄ°RLEÅTÄ°R =====
        # EÄŸer hem hareketler hem arÅŸivde aynÄ± sÃ¼tun varsa (Ã¶rn: tedarikci)
        # Hareketlerdeki boÅŸsa arÅŸivden doldur
        for col in ['tedarikci', 'yore', 'bugday_cinsi']:
            if col in merged.columns and f'{col}_arsiv' in merged.columns:
                merged[col] = merged[col].fillna(merged[f'{col}_arsiv'])
                # Gereksiz _arsiv sÃ¼tununu sil
                merged.drop(f'{col}_arsiv', axis=1, inplace=True)
        
        # ===== TARÄ°H SIRALAMASI =====
        if 'tarih' in merged.columns:
            merged['tarih'] = pd.to_datetime(merged['tarih'], errors='coerce')
            merged = merged.sort_values('tarih', ascending=False)
        
        return merged
        
    except Exception as e:
        st.error(f"âŒ Hareket yÃ¼kleme hatasÄ±: {e}")
        import traceback
        st.code(traceback.format_exc())
        return pd.DataFrame()
        
        # Ã‡akÄ±ÅŸan sÃ¼tunlarda boÅŸluklarÄ± doldur
        for col in ['tedarikci', 'yore']:
            if f'{col}_arsiv' in merged.columns:
                merged[col] = merged[col].fillna(merged[f'{col}_arsiv'])
        
        if 'tarih' in merged.columns:
            merged['tarih'] = pd.to_datetime(merged['tarih'])
            merged = merged.sort_values('tarih', ascending=False)
            
        return merged
    except Exception as e:
        st.error(f"Hareket yÃ¼kleme hatasÄ±: {e}")
        return pd.DataFrame()

def get_bugday_arsiv():
    """ArÅŸiv verisi"""
    df = fetch_data("bugday_giris_arsivi")
    if not df.empty and 'tarih' in df.columns:
        df['tarih'] = pd.to_datetime(df['tarih'])
        df = df.sort_values('tarih', ascending=False)
    return df

# --- TAVLI ANALÄ°ZLERÄ° (TEMPERED WHEAT) ---

def save_tavli_analiz(silo_isim, analiz_tonaj, **analiz_degerleri):
    try:
        data = {
            'silo_isim': silo_isim, 
            'analiz_tonaj': float(analiz_tonaj),
            'tarih': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            **analiz_degerleri
        }
        return add_data("tavli_analiz", data), "Kaydedildi"
    except Exception as e: return False, str(e)

def get_tavli_analizler(silo_isim=None):
    df = fetch_data("tavli_analiz")
    if df.empty: return pd.DataFrame()
    if silo_isim: df = df[df['silo_isim'] == silo_isim]
    if 'tarih' in df.columns:
        df['tarih'] = pd.to_datetime(df['tarih'])
        df = df.sort_values('tarih', ascending=False)
    return df
def get_kuru_bugday_agirlikli_ortalama(silo_isim):
    """
    Bir silodaki KURU BUÄDAY analizlerinin aÄŸÄ±rlÄ±klÄ± ortalamasÄ±nÄ± hesaplar.
    Mal kabul giriÅŸlerinden (hareketler tablosu) veriler alÄ±nÄ±r.
    
    Returns:
        dict: AÄŸÄ±rlÄ±klÄ± ortalama analiz deÄŸerleri
    """
    try:
        # Hareketler tablosundan bu silonun GÄ°RÄ°Å kayÄ±tlarÄ±nÄ± al
        df_hareketler = fetch_data("hareketler")
        if df_hareketler.empty:
            return {}
        
        # Sadece bu silonun giriÅŸleri
        df_silo = df_hareketler[
            (df_hareketler['silo_isim'] == silo_isim) & 
            (df_hareketler['hareket_tipi'] == 'GiriÅŸ')
        ].copy()
        
        if df_silo.empty:
            return {}
        
        # Tonaj sÃ¼tunu kontrolÃ¼
        if 'miktar' not in df_silo.columns:
            return {}
        
        # Numeric dÃ¶nÃ¼ÅŸÃ¼m
        numeric_cols = ['miktar', 'hektolitre', 'protein', 'rutubet', 'gluten', 
                       'gluten_index', 'sedim', 'gecikmeli_sedim']
        
        for col in numeric_cols:
            if col in df_silo.columns:
                df_silo[col] = pd.to_numeric(df_silo[col], errors='coerce').fillna(0)
        
        toplam_tonaj = df_silo['miktar'].sum()
        
        if toplam_tonaj == 0:
            return {}
        
        # AÄŸÄ±rlÄ±klÄ± ortalama hesapla
        ortalama = {}
        analiz_cols = ['hektolitre', 'protein', 'rutubet', 'gluten', 
                      'gluten_index', 'sedim', 'gecikmeli_sedim']
        
        for col in analiz_cols:
            if col in df_silo.columns:
                # (miktar * deÄŸer).sum() / toplam_miktar
                agirlikli_toplam = (df_silo['miktar'] * df_silo[col]).sum()
                ortalama[col] = agirlikli_toplam / toplam_tonaj if toplam_tonaj > 0 else 0
        
        return ortalama
        
    except Exception as e:
        st.error(f"Kuru buÄŸday ortalama hesaplama hatasÄ±: {e}")
        return {}
# --- SPEC YÃ–NETÄ°MÄ° ---

def save_bugday_spec(bugday_cinsi, parametre, min_val, max_val, hedef_val):
    try:
        conn = get_conn()
        df = fetch_data("bugday_spekleri")
        new_row = {
            'bugday_cinsi': bugday_cinsi, 'parametre': parametre, 
            'min_deger': min_val, 'max_deger': max_val, 'hedef_deger': hedef_val, 'aktif': 1
        }
        
        if df.empty: return add_data("bugday_spekleri", new_row)
        
        # Upsert Logic
        mask = (df['bugday_cinsi'] == bugday_cinsi) & (df['parametre'] == parametre)
        if mask.any():
            df.loc[mask, ['min_deger', 'max_deger', 'hedef_deger']] = [min_val, max_val, hedef_val]
            conn.update(worksheet="bugday_spekleri", data=df)
        else:
            add_data("bugday_spekleri", new_row)
        return True
    except: return False

def get_all_bugday_specs_dataframe():
    df = fetch_data("bugday_spekleri")
    return df if not df.empty else pd.DataFrame()

def delete_bugday_spec_group(cins):
    try:
        conn = get_conn()
        df = fetch_data("bugday_spekleri")
        if df.empty: return True
        df = df[df['bugday_cinsi'] != cins]
        conn.update(worksheet="bugday_spekleri", data=df)
        return True
    except: return False

# --------------------------------------------------------------------------
# UI EKRANLARI - %100 ORÄ°JÄ°NAL KAPSAM (EKSÄ°KSÄ°Z)
# --------------------------------------------------------------------------

def show_mal_kabul():
    """Mal Kabul EkranÄ± - Dinamik Lot ve Profesyonel Terminoloji"""
    if st.session_state.get('user_role') not in ["admin", "operations", "quality"]:
        st.warning("Yetkisiz EriÅŸim")
        return

    st.header(f"ğŸšœ {t('header_goods_receipt')}")
    
    # --- DÄ°NAMÄ°K LOT NUMARASI MANTIÄI ---
    # SeÃ§ili dile gÃ¶re Lot Ã¶nekini (Prefix) deÄŸiÅŸtiriyoruz
    lang_code = st.session_state.get('language_code', 'TR')
    
    prefix_map = {
        'TR': 'BUGDAY',
        'EN': 'WHEAT',
        'FR': 'BLE',
        'RU': 'PSHN' # veya 'ĞŸĞ¨Ğ' (Kiril karakter bazen sorun olabilir ama deneyebiliriz)
    }
    lot_prefix = prefix_map.get(lang_code, 'BUGDAY')
    
    # Lot numarasÄ±nÄ± oluÅŸtur
    timestamp = datetime.now().strftime('%y%m%d%H%M%S')
    lot_no = f"{lot_prefix}-{timestamp}"
    
    col1, col2 = st.columns([1, 1.5], gap="large")
    
    with col1:
        st.subheader(f"ğŸ“‹ {t('subheader_basic_info')}")
        
        # Dinamik etiketli Lot No
        st.info(f"**{t('label_lot')}:** `{lot_no}`")
        
        
        df_silo = get_silo_data()
        if df_silo.empty: 
            st.warning("Silo tanÄ±mlayÄ±nÄ±z.")
            return
            
        secilen_silo = st.selectbox(f"{t('label_silo')} *", df_silo['isim'].tolist())
        
        # Kapasite
        silo_row = df_silo[df_silo['isim'] == secilen_silo].iloc[0]
        mevcut = float(silo_row.get('mevcut_miktar', 0))
        kapasite = float(silo_row.get('kapasite', 0))
        kalan = kapasite - mevcut
        st.info(f"{t('label_balance')}: {kalan:.1f} Ton")
        
        tarih = st.date_input(f"{t('label_date')} *", datetime.now())
        
        # Standart SeÃ§imi
        specs_list = []
        df_specs = fetch_data("bugday_spekleri")
        if not df_specs.empty:
            specs_list = df_specs['bugday_cinsi'].unique().tolist()
            
        secilen_standart = st.selectbox(t("label_standard"), ["( - )"] + specs_list)
        
        bugday_cinsi = st.text_input(f"{t('label_variety')} *")
        
        current_specs = {}
        if secilen_standart != "( - )":
            df_s = df_specs[df_specs['bugday_cinsi'] == secilen_standart]
            for _, row in df_s.iterrows():
                current_specs[row['parametre']] = row

        tedarikci = st.text_input(f"{t('label_supplier')} *")
        yore = st.text_input(f"{t('label_origin')} *")
        plaka = st.text_input(f"{t('label_plate')} *")
        notlar = st.text_area(t("label_notes"), key="mal_kabul_notlar")
        
        miktar = st.number_input(f"{t('label_weight')} *", min_value=0.0, format="%.1f")
        fiyat = st.number_input(f"{t('label_price')} *", min_value=0.0, format="%.2f")

    with col2:
        st.subheader(f"ğŸ§ª {t('subheader_quality')}")
        
        # Validasyon Helper
        def validate_val(key, val, label):
            if key in current_specs:
                spec = current_specs[key]
                s_min = float(spec.get('min_deger', 0))
                s_max = float(spec.get('max_deger', 999))
                if val < s_min or (s_max > 0 and val > s_max):
                    st.error(f"âŒ {label} Limit DÄ±ÅŸÄ±!")

        c1, c2, c3 = st.columns(3)
        
        with c1:
            # Hektolitre (Test Weight)
            g_hl = st.number_input(t("ana_test_weight"), 0.0, 100.0, 78.0)
            validate_val("hektolitre", g_hl, "HL")
            
            # Rutubet (Moisture)
            g_rut = st.number_input(t("ana_moisture"), 0.0, 20.0, 13.5)
            validate_val("rutubet", g_rut, "Rutubet")
            
            # Protein
            g_prot = st.number_input(t("ana_protein"), 0.0, 20.0, 12.0)
            validate_val("protein", g_prot, "Protein")
            
            # Gluten (Wet Gluten)
            g_glut = st.number_input(t("ana_gluten"), 0.0, 50.0, 28.0)
            validate_val("gluten", g_glut, "Gluten")

        with c2:
            # Gluten Index
            g_index = st.number_input(t("ana_gluten_index"), 0.0, 100.0, 90.0)
            validate_val("gluten_index", g_index, "G.Index")
            
            # Sedim (Zeleny)
            g_sedim = st.number_input(t("ana_sedim"), 0.0, 100.0, 30.0)
            validate_val("sedim", g_sedim, "Sedim")
            
            # Gecikmeli Sedim (henÃ¼z excelde yoktu, sabit kalabilir veya eklenebilir)
            g_g_sedim = st.number_input("G.Sedim / Delayed", 0.0, 100.0, 35.0)
            
            # SÃ¼ne
            sune = st.number_input("SÃ¼ne (%)", 0.0, 10.0, 0.5)

        with c3:
            kirik_ciliz = st.number_input("KÄ±rÄ±k/Broken (%)", 0.0, 100.0, 3.0)
            yabanci_tane = st.number_input("YabancÄ±/Foreign (%)", 0.0, 100.0, 3.5)
            hasere = st.selectbox("HaÅŸere/Insect", ["Yok/None", "Var/Yes"])

    st.divider()
    
    # Kaydet Butonu
    if st.button(f"ğŸ’¾ {t('btn_submit')}", type="primary", use_container_width=True):
        # ... (Validasyon kodlarÄ± aynen kalÄ±yor, buraya dokunma) ...
        # ... (Validasyon kodlarÄ± aynen kalÄ±yor, buraya dokunma) ...

        # ğŸ‘‡ SPINNER BURADA BAÅLIYOR ğŸ‘‡
        with st.spinner("VeritabanÄ±na kaydediliyor, lÃ¼tfen bekleyiniz..."):
            
            # 1. Stok Hareketi Logla
            ok_log = log_stok_hareketi(
                secilen_silo, "GiriÅŸ", miktar,
                protein=g_prot, gluten=g_glut, rutubet=g_rut, hektolitre=g_hl,
                sedim=g_sedim, maliyet=fiyat, lot_no=lot_no,
                tedarikci=tedarikci, yore=yore, notlar=note_final
            )
            
            if ok_log:
                # 2. ArÅŸive Ekle
                ok_arc = add_to_bugday_giris_arsivi(
                    lot_no, tarih=str(tarih), bugday_cinsi=bugday_cinsi,
                    tedarikci=tedarikci, yore=yore, plaka=plaka,
                    tonaj=miktar, fiyat=fiyat, silo_isim=secilen_silo,
                    hektolitre=g_hl, protein=g_prot, rutubet=g_rut,
                    gluten=g_glut, gluten_index=g_index, sedim=g_sedim,
                    gecikmeli_sedim=g_g_sedim, sune=sune, kirik_ciliz=kirik_ciliz,
                    yabanci_tane=yabanci_tane, notlar=note_final
                )
                
                if ok_arc:
                    st.success("âœ… KayÄ±t BaÅŸarÄ±lÄ±!")
                    recalculate_silos_from_logs() # SilolarÄ± gÃ¼ncelle (Bu iÅŸlem uzun sÃ¼rer)
                    time.sleep(1)
                    st.rerun()
                else:
                    st.error("ArÅŸiv kaydÄ±nda hata oluÅŸtu.")
            else:
                st.error("Stok kaydÄ±nda hata oluÅŸtu.")

def show_stok_cikis():
    """Stok Ã‡Ä±kÄ±ÅŸÄ± EkranÄ± - AKILLI TRANSFER VE KALÄ°TE TAÅIMA Ã–ZELLÄ°KLÄ°"""
    st.header("ğŸ“‰ Stok Ã‡Ä±kÄ±ÅŸÄ± (Ãœretim/Transfer)")
    
    # Verileri Ã‡ek
    df = get_silo_data()
    if df.empty: 
        st.warning("Silo bulunamadÄ±.")
        return
    
    col1, col2 = st.columns(2)
    with col1:
        silo = st.selectbox("Kaynak Silo", df['isim'].tolist())
        row = df[df['isim'] == silo].iloc[0]
        mevcut = float(row['mevcut_miktar'])
        st.metric("Mevcut", f"{mevcut:.1f} Ton")
        
        miktar = st.number_input("Miktar (Ton)", 0.1, max_value=mevcut if mevcut > 0 else 0.1)
        neden = st.selectbox("Neden", ["Ãœretime GÃ¶nderim", "Silo Transferi", "SatÄ±ÅŸ", "Zayi"])
        
        hedef = None
        if neden == "Silo Transferi":
            hedef = st.selectbox("Hedef Silo", [s for s in df['isim'].tolist() if s != silo])
            
    with col2:
        # GÃ¶rsel Ã–nizleme
        yeni = max(0, mevcut - miktar)
        doluluk = yeni / float(row['kapasite']) if float(row['kapasite']) > 0 else 0
        st.markdown(draw_silo(doluluk, f"Kalan: {yeni:.1f}"), unsafe_allow_html=True)
    
    st.divider()
    
    if st.button("ğŸ“¤ Ã‡Ä±kÄ±ÅŸÄ± Onayla", type="primary", use_container_width=True):
        # ===== 1. VALÄ°DASYONLAR =====
        from app.core.config import validate_stock_withdrawal, validate_capacity
        
        validasyon_hatalari = []
        
        # Stok Yeterlilik KontrolÃ¼
        valid, msg = validate_stock_withdrawal(mevcut, miktar)
        if not valid: validasyon_hatalari.append(msg)
        
        # Transfer Hedef KontrolÃ¼
        if neden == "Silo Transferi":
            if not hedef:
                validasyon_hatalari.append("âŒ Hedef silo seÃ§melisiniz!")
            else:
                hedef_row = df[df['isim'] == hedef].iloc[0]
                hedef_mevcut = float(hedef_row['mevcut_miktar'])
                hedef_kapasite = float(hedef_row['kapasite'])
                valid_cap, msg_cap, _ = validate_capacity(hedef_mevcut, hedef_kapasite, miktar)
                if not valid_cap: validasyon_hatalari.append(f"Hedef Silo: {msg_cap}")
        
        if validasyon_hatalari:
            st.error("ğŸš« Hata:")
            for hata in validasyon_hatalari: st.write(f"- {hata}")
            return
        
        # ===== 2. Ä°ÅLEM BAÅLIYOR =====
        
        # A) Kaynak Silodan Ã‡Ä±kÄ±ÅŸ (Standart Ä°ÅŸlem - Google Sheets 'hareketler' tablosuna yazar)
        if log_stok_hareketi(silo, "Ã‡Ä±kÄ±ÅŸ", miktar, notlar=neden):
            update_tavli_bugday_stok(silo, miktar, "cikar")
            
            # B) TRANSFER Ä°SE: AKILLI KALÄ°TE KOPYALAMA
            if neden == "Silo Transferi" and hedef:
                try:
                    # 1. Kaynak Silonun Kalite DNA'sÄ±nÄ± Ã‡Ä±kar (Mixing ModÃ¼lÃ¼nden)
                    # Bu fonksiyon Google Sheets'teki 'tavli_analiz' tablosunu okuyup aÄŸÄ±rlÄ±klÄ± ortalamayÄ± hesaplar.
                    from app.modules.mixing import get_tavli_analiz_agirlikli_ortalama
                    
                    kaynak_analiz = get_tavli_analiz_agirlikli_ortalama(silo)
                    
                    if not kaynak_analiz:
                        # EÄŸer detaylÄ± analiz yoksa, en azÄ±ndan temel bilgileri silolar tablosundan al
                        kaynak_analiz = {
                            'protein': float(row.get('protein', 0)),
                            'gluten': float(row.get('gluten', 0)),
                            'rutubet': float(row.get('rutubet', 0)),
                            'hektolitre': float(row.get('hektolitre', 0)),
                            'sedim': float(row.get('sedim', 0)),
                            'maliyet': float(row.get('maliyet', 0))
                        }
                    
                    # Gereksiz sistem alanlarÄ±nÄ± temizle
                    if 'toplam_tonaj' in kaynak_analiz: del kaynak_analiz['toplam_tonaj']
                    if 'analiz_sayisi' in kaynak_analiz: del kaynak_analiz['analiz_sayisi']
                    
                    # 2. Hedef Siloya "GiriÅŸ" Hareketi Yaz (Google Sheets 'hareketler' tablosu)
                    log_stok_hareketi(
                        hedef, 
                        "GiriÅŸ", 
                        miktar, 
                        # Hesaplanan ortalamalarÄ± buraya aktarÄ±yoruz
                        protein=kaynak_analiz.get('protein', 0),
                        gluten=kaynak_analiz.get('gluten', 0),
                        rutubet=kaynak_analiz.get('rutubet', 0),
                        hektolitre=kaynak_analiz.get('hektolitre', 0),
                        sedim=kaynak_analiz.get('sedim', 0),
                        maliyet=kaynak_analiz.get('maliyet', 0),
                        notlar=f"Transfer: {silo} -> {hedef}"
                    )
                    
                    # 3. Hedef Siloya "TavlÄ± Analiz" KaydÄ± Yaz (Google Sheets 'tavli_analiz' tablosu)
                    # BurasÄ± Farino/Extenso gibi detaylÄ± verilerin taÅŸÄ±ndÄ±ÄŸÄ± kritik nokta!
                    save_tavli_analiz(
                        hedef, 
                        miktar, # Transfer edilen tonaj kadar aÄŸÄ±rlÄ±ÄŸÄ± olur
                        **kaynak_analiz, # TÃ¼m analiz parametrelerini aÃ§Ä±p kaydediyoruz
                        notlar=f"Transfer Kaynak: {silo}"
                    )
                    
                    # 4. Hedef Silonun TavlÄ± StoÄŸunu ArtÄ±r (Google Sheets 'silolar' tablosu)
                    update_tavli_bugday_stok(hedef, miktar, "ekle")
                    
                    st.success(f"âœ… {silo} silosundaki kalite deÄŸerleri {hedef} silosuna {miktar} ton aÄŸÄ±rlÄ±kla iÅŸlendi.")
                    
                except Exception as e:
                    st.error(f"Transfer analiz taÅŸÄ±ma hatasÄ±: {e}")
            
            # C) TÃ¼m SilolarÄ± Yeniden Hesapla (Google Sheets Senkronizasyonu)
            # Bu fonksiyon 'hareketler' tablosunu okuyup 'silolar' tablosundaki gÃ¼ncel stok ve ortalamalarÄ± dÃ¼zeltir.
            recalculate_silos_from_logs()
            
            st.success("âœ… Stok ve Analiz Transferi TamamlandÄ±!")
            time.sleep(1)
            st.rerun()
        else:
            st.error("âŒ Ã‡Ä±kÄ±ÅŸ kaydÄ± oluÅŸturulamadÄ±!")
def show_tavli_analiz():
    """TavlÄ± BuÄŸday Analizi - TAM VE EKSÄ°KSÄ°Z Parametreler"""
    st.header("ğŸ§ª TavlÄ± BuÄŸday Analiz KaydÄ±")
    df = get_silo_data()
    if df.empty: 
        st.warning("Silo bulunamadÄ±")
        return
    
    c1, c2 = st.columns(2)
    with c1:
        silo = st.selectbox("Silo SeÃ§", df['isim'].tolist())
        row = df[df['isim'] == silo].iloc[0]
        mevcut = float(row.get('mevcut_miktar', 0))
        
        # TavlÄ± stok kontrolÃ¼ - SÃ¼tun adÄ±nÄ± kontrol et
        tavli_col = 'tavli_bugday_stok' if 'tavli_bugday_stok' in df.columns else 'tavli_stok'
        tavli = float(row.get(tavli_col, 0)) if pd.notnull(row.get(tavli_col, 0)) else 0.0
        
        kalan = max(0, mevcut - tavli)
        st.info(f"Mevcut: {mevcut:.1f} | TavlÄ±: {tavli:.1f} | Eklenebilir: {kalan:.1f}")
        
        tonaj = st.number_input("Analiz TonajÄ±", 0.1, max_value=max(kalan, 1000.0), value=min(kalan, 10.0) if kalan > 0 else 10.0)
    
    with c2:
        tarih = st.date_input("Tarih", datetime.now())
        notlar = st.text_area("Notlar", key="tavli_notlar")

    # Tabs - TAM VERSÄ°YON
    tab1, tab2, tab3 = st.tabs(["ğŸ§ª Kimyasal", "ğŸ“ˆ Farinograph", "ğŸ“Š Extensograph"])
    vals = {}
    
    with tab1:
        cc1, cc2 = st.columns(2)
        vals['protein'] = cc1.number_input("Protein (%)", value=float(row.get('protein', 12.0)), format="%.2f")
        vals['rutubet'] = cc1.number_input("Rutubet (%)", value=15.0, format="%.2f")
        vals['gluten'] = cc1.number_input("Gluten (%)", value=float(row.get('gluten', 28.0)), format="%.2f")
        vals['gluten_index'] = cc1.number_input("Gluten Index", value=95.0, format="%.2f")
        
        vals['sedim'] = cc2.number_input("Sedim (ml)", value=50.0, format="%.2f")
        vals['g_sedim'] = cc2.number_input("G. Sedim (ml)", value=60.0, format="%.2f")
        vals['fn'] = cc2.number_input("FN", value=300.0, format="%.2f")
        vals['ffn'] = cc2.number_input("FFN", value=400.0, format="%.2f")
        vals['amilograph'] = cc2.number_input("Amilograph", value=1100.0, format="%.2f")
        
    with tab2:
        cc1, cc2 = st.columns(2)
        vals['su_kaldirma_f'] = cc1.number_input("Su KaldÄ±rma (Farino) (%)", value=58.0, format="%.2f")
        vals['gelisme_suresi'] = cc1.number_input("GeliÅŸme SÃ¼resi (dk)", value=3.0, format="%.2f")
        vals['stabilite'] = cc2.number_input("Stabilite (dk)", value=8.0, format="%.2f")
        vals['yumusama'] = cc2.number_input("YumuÅŸama (FU)", value=70.0, format="%.2f")
        
    with tab3:
        st.subheader("ğŸ“Š Extensograph Analizleri (DetaylÄ±)")
        vals['su_kaldirma_e'] = st.number_input("Su KaldÄ±rma (Extenso) (%)", value=58.0, format="%.2f")
        
        # 45 DAKÄ°KA
        with st.expander("ğŸ“Š 45. Dakika:", expanded=True):
            cols45 = st.columns(3)
            vals['direnc45'] = cols45[0].number_input("DirenÃ§ (45)", value=610.0, format="%.2f", key="d45")
            vals['taban45'] = cols45[1].number_input("Taban (45)", value=165.0, format="%.2f", key="t45")
            vals['enerji45'] = cols45[2].number_input("Enerji (45)", value=110.0, format="%.2f", key="e45")
        
        # 90 DAKÄ°KA
        with st.expander("ğŸ“Š 90. Dakika:", expanded=True):
            cols90 = st.columns(3)
            vals['direnc90'] = cols90[0].number_input("DirenÃ§ (90)", value=900.0, format="%.2f", key="d90")
            vals['taban90'] = cols90[1].number_input("Taban (90)", value=125.0, format="%.2f", key="t90")
            vals['enerji90'] = cols90[2].number_input("Enerji (90)", value=120.0, format="%.2f", key="e90")
        
        # 135 DAKÄ°KA
        with st.expander("ğŸ“Š 135. Dakika:", expanded=True):
            cols135 = st.columns(3)
            vals['direnc135'] = cols135[0].number_input("DirenÃ§ (135)", value=980.0, format="%.2f", key="d135")
            vals['taban135'] = cols135[1].number_input("Taban (135)", value=120.0, format="%.2f", key="t135")
            vals['enerji135'] = cols135[2].number_input("Enerji (135)", value=126.0, format="%.2f", key="e135")

    st.divider()
    if st.button("ğŸ’¾ Kaydet", type="primary", use_container_width=True):
        if tonaj > kalan + 0.1:
            st.error(f"âŒ Kapasite hatasÄ±: Sadece {kalan:.1f} ton eklenebilir!")
            return
        
        # 1. TavlÄ± analiz kaydet
        ok, msg = save_tavli_analiz(silo, tonaj, **vals, notlar=notlar, tarih=str(tarih))
        
        if ok:
            # 2. TavlÄ± stoku gÃ¼ncelle - DÃœZELTÄ°LMÄ°Å VERSÄ°YON
            try:
                conn = get_conn()
                df_update = fetch_data("silolar")
                
                # DEBUG: Mevcut sÃ¼tunlarÄ± gÃ¶ster
                st.info(f"ğŸ“Š Silolar tablosundaki sÃ¼tunlar: {list(df_update.columns)}")
                
                if not df_update.empty:
                    mask = df_update['isim'] == silo
                    
                    if mask.any():
                        # SÃ¼tun adÄ±nÄ± kontrol et - TÃœM OLASILIKLARÄ± KAPSAYAN VERSÄ°YON
                        tavli_col = None
                        for col_name in ['tavli_bugday_stok', 'tavli_stok', 'tavli_bugday', 'tavlÄ±_stok']:
                            if col_name in df_update.columns:
                                tavli_col = col_name
                                break
                        
                        # EÄŸer sÃ¼tun yoksa oluÅŸtur
                        if tavli_col is None:
                            st.warning("âš ï¸ TavlÄ± stok sÃ¼tunu bulunamadÄ±, 'tavli_bugday_stok' oluÅŸturuluyor...")
                            df_update['tavli_bugday_stok'] = 0.0
                            tavli_col = 'tavli_bugday_stok'
                        
                        st.info(f"ğŸ” KullanÄ±lan sÃ¼tun adÄ±: **{tavli_col}**")
                        
                        # Mevcut tavlÄ± stoku al
                        current_tavli = float(df_update.loc[mask, tavli_col].iloc[0]) if pd.notnull(df_update.loc[mask, tavli_col].iloc[0]) else 0.0
                        
                        # Yeni tavlÄ± stok hesapla
                        yeni_tavli = current_tavli + float(tonaj)
                        
                        # GÃ¼ncelle
                        df_update.loc[mask, tavli_col] = yeni_tavli
                        conn.update(worksheet="silolar", data=df_update)
                        
                        st.success(f"âœ… TavlÄ± analiz kaydedildi! TavlÄ± Stok: {current_tavli:.1f} â†’ {yeni_tavli:.1f} Ton")
                        time.sleep(2)
                        st.rerun()
                    else:
                        st.error("Silo bulunamadÄ±!")
                else:
                    st.error("Silo verisi yÃ¼klenemedi!")
                    
            except Exception as e:
                st.error(f"âŒ Stok gÃ¼ncelleme hatasÄ±: {str(e)}")
                st.error(f"ğŸ” Debug: {type(e).__name__}")
        else:
            st.error(f"âŒ KayÄ±t hatasÄ±: {msg}")
def show_stok_hareketleri():
    """
    Stok Hareketleri (Dijital Defter) ModÃ¼lÃ¼
    GiriÅŸler YEÅÄ°L, Ã‡Ä±kÄ±ÅŸlar KIRMIZI olarak listelenir.
    """
    st.header("ğŸ”„ Stok Hareket Defteri")
    st.info("BurasÄ± silolara giren (+) ve silolardan Ã¼retime/satÄ±ÅŸa Ã§Ä±kan (-) tÃ¼m hareketlerin Ã¶zetidir.")
    
    # Veriyi Ã‡ek
    df = fetch_data("hareketler")
    
    if df.empty:
        st.warning("ğŸ“­ HenÃ¼z stok hareket kaydÄ± bulunmamaktadÄ±r.")
        return

    # Tarih formatÄ± dÃ¼zenleme
    if 'tarih' in df.columns:
        df['tarih'] = pd.to_datetime(df['tarih'], errors='coerce')
        df = df.sort_values('tarih', ascending=False)

    # --- FÄ°LTRELER ---
    col_f1, col_f2, col_f3 = st.columns(3)
    
    with col_f1:
        # Silo Filtresi
        silolar = ["TÃ¼mÃ¼"] + sorted(df['silo_isim'].dropna().unique().tolist()) if 'silo_isim' in df.columns else ["TÃ¼mÃ¼"]
        secilen_silo = st.selectbox("Depo/Silo SeÃ§:", silolar, key="filtre_silo_hareket")
        
    with col_f2:
        # Ä°ÅŸlem TÃ¼rÃ¼ Filtresi
        turler = ["TÃ¼mÃ¼", "GiriÅŸ", "Ã‡Ä±kÄ±ÅŸ", "Transfer"]
        secilen_tur = st.selectbox("Ä°ÅŸlem TÃ¼rÃ¼:", turler, key="filtre_tur_hareket")
        
    with col_f3:
        # Tarih AralÄ±ÄŸÄ±
        bugun = datetime.now().date()
        baslangic = st.date_input("BaÅŸlangÄ±Ã§", bugun - pd.Timedelta(days=30), key="filtre_tarih_bas")
    
    # Filtreleri Uygula
    filtered_df = df.copy()
    
    if secilen_silo != "TÃ¼mÃ¼":
        filtered_df = filtered_df[filtered_df['silo_isim'] == secilen_silo]
        
    if secilen_tur != "TÃ¼mÃ¼":
        filtered_df = filtered_df[filtered_df['hareket_tipi'].astype(str).str.contains(secilen_tur, case=False, na=False)]
        
    if 'tarih' in filtered_df.columns:
        filtered_df = filtered_df[filtered_df['tarih'].dt.date >= baslangic]

    # --- TABLO GÃ–RÃœNÃœMÃœ ---
    st.write(f"ğŸ“‹ **Toplam KayÄ±t:** {len(filtered_df)}")
    
    if not filtered_df.empty:
        # GÃ¶rÃ¼ntÃ¼lenecek SÃ¼tunlar
        display_cols = ['tarih', 'hareket_tipi', 'silo_isim', 'miktar', 'lot_no', 'notlar']
        # Sadece var olanlarÄ± al
        display_cols = [c for c in display_cols if c in filtered_df.columns]
        
        df_view = filtered_df[display_cols].copy()
        
        # Kolon Ä°simlerini TÃ¼rkÃ§eleÅŸtir
        col_map = {
            'tarih': 'Tarih',
            'hareket_tipi': 'Ä°ÅŸlem',
            'silo_isim': 'Silo',
            'miktar': 'Miktar (Ton)',
            'lot_no': 'Lot No',
            'notlar': 'AÃ§Ä±klama'
        }
        df_view = df_view.rename(columns=col_map)
        
        # Renklendirme Fonksiyonu
        def highlight_rows(row):
            islem = str(row.get('Ä°ÅŸlem', '')).lower()
            if 'giriÅŸ' in islem or 'giris' in islem:
                return ['background-color: #dcfce7; color: #14532d'] * len(row)  # AÃ§Ä±k YeÅŸil
            elif 'Ã§Ä±kÄ±ÅŸ' in islem or 'cikis' in islem:
                return ['background-color: #fee2e2; color: #7f1d1d'] * len(row)  # AÃ§Ä±k KÄ±rmÄ±zÄ±
            elif 'transfer' in islem:
                return ['background-color: #e0f2fe; color: #0c4a6e'] * len(row)  # AÃ§Ä±k Mavi
            return [''] * len(row)

        st.dataframe(
            df_view.style.apply(highlight_rows, axis=1).format({"Miktar (Ton)": "{:.1f}", "Tarih": lambda t: t.strftime("%d.%m.%Y %H:%M")}),
            use_container_width=True,
            height=600
        )
    else:
        st.info("Kriterlere uygun kayÄ±t bulunamadÄ±.")

# ==============================================================================
# Ã–ZEL EXCEL RAPOR MOTORU (BUÄDAY GÄ°RÄ°Å - GRUPLANDIRILMIÅ BAÅLIKLAR)
# ==============================================================================
def export_bugday_giris_ozel_excel(df):
    """
    BuÄŸday GiriÅŸ ArÅŸivi iÃ§in Ã¶zel gruplandÄ±rÄ±lmÄ±ÅŸ Excel Ã¼retir.
    YapÄ±: [TEMEL BÄ°LGÄ°LER] + [FÄ°ZÄ°KSEL VE KÄ°MYASAL ANALÄ°ZLER]
    """
    try:
        from io import BytesIO
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = "Bugday Giris Raporu"

        # HaÅŸere Durumu Hesapla (Excel iÃ§in)
        if 'notlar' in df.columns:
            df['Hasere_Durum'] = df['notlar'].apply(
                lambda x: "VAR" if isinstance(x, str) and any(k in x.upper() for k in ["HAÅ", "BÃ–C", "BIT", "CANLI"]) else "TEMÄ°Z"
            )
        else:
            df['Hasere_Durum'] = "TEMÄ°Z"

        # --- TASARIM TANIMLARI ---
        structure = [
            {
                "group": "TEMEL BÄ°LGÄ°LER",
                "color": "4472C4", # Mavi
                "cols": [
                    ("ID No", "id_counter"), # Ã–zel oluÅŸturulacak
                    ("Tarih", "tarih"),
                    ("Lot No", "lot_no"),
                    ("Plaka", "plaka"),
                    ("BuÄŸday Cinsi", "bugday_cinsi"),
                    ("TedarikÃ§i AdÄ±", "tedarikci"),
                    ("YÃ¶re/BÃ¶lge", "yore"),
                    ("DÃ¶kÃ¼ldÃ¼ÄŸÃ¼ Silo", "silo_isim"),
                    ("Tonaj", "tonaj"),
                    ("Fiyat", "fiyat")
                ]
            },
            {
                "group": "FÄ°ZÄ°KSEL VE KÄ°MYASAL ANALÄ°ZLER",
                "color": "ED7D31", # Turuncu
                "cols": [
                    ("Hektolitre", "hektolitre"),
                    ("Protein", "protein"),
                    ("Rutubet", "rutubet"),
                    ("Gluten", "gluten"),
                    ("Gluten Ä°ndeks", "gluten_index"),
                    ("Sedim", "sedim"),
                    ("G. Sedim", "gecikmeli_sedim"),
                    ("SÃ¼ne", "sune"),
                    ("KÄ±rÄ±k & CÄ±lÄ±z", "kirik_ciliz"),
                    ("YabancÄ± Tane", "yabanci_tane"),
                    ("HaÅŸere", "Hasere_Durum"), # Hesaplanan sÃ¼tun
                    ("Not", "notlar")
                ]
            }
        ]

        # --- STÄ°LLER ---
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        header_font = Font(bold=True, color="FFFFFF", size=11)
        sub_header_font = Font(bold=True, color="000000", size=10)
        
        # --- BAÅLIKLARI YAZMA ---
        current_col = 1
        for group in structure:
            start_col = current_col
            num_cols = len(group["cols"])
            end_col = start_col + num_cols - 1
            
            # 1. Ãœst BaÅŸlÄ±k (Merge)
            ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
            cell = ws.cell(row=1, column=start_col, value=group["group"])
            cell.fill = PatternFill("solid", fgColor=group["color"])
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
            
            for c in range(start_col, end_col + 1):
                ws.cell(row=1, column=c).border = thin_border

            # 2. Alt BaÅŸlÄ±klar
            for i, (col_name, db_key) in enumerate(group["cols"]):
                cell_sub = ws.cell(row=2, column=start_col + i, value=col_name)
                cell_sub.font = sub_header_font
                cell_sub.alignment = Alignment(horizontal="center", vertical="center")
                cell_sub.border = thin_border
                cell_sub.fill = PatternFill("solid", fgColor="E7E6E6")

            current_col += num_cols

        # --- VERÄ°LERÄ° YAZMA ---
        records = df.to_dict('records')
        for r_idx, row_data in enumerate(records, start=3):
            current_col = 1
            for group in structure:
                for col_name, db_key in group["cols"]:
                    
                    # Ã–zel Durumlar
                    if db_key == "id_counter":
                        val = r_idx - 2 # 1'den baÅŸlat
                    else:
                        val = row_data.get(db_key, "")
                    
                    # Tarih FormatÄ±
                    if db_key == "tarih" and val:
                        try: val = pd.to_datetime(val).strftime('%d.%m.%Y')
                        except: pass
                    
                    # SayÄ±sal Yuvarlama
                    try:
                        if isinstance(val, (int, float)):
                            val = round(float(val), 1)
                        elif val and db_key not in ["tarih", "silo_isim", "notlar", "plaka", "Hasere_Durum", "tedarikci", "yore", "bugday_cinsi", "lot_no"]:
                            val = round(float(val), 1)
                    except: pass

                    cell = ws.cell(row=r_idx, column=current_col, value=val)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal="center")
                    current_col += 1

        # --- SÃœTUN GENÄ°ÅLÄ°KLERÄ° ---
        for i, col in enumerate(ws.columns, 1):
            max_length = 0
            column_letter = get_column_letter(i)
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except: pass
            adjusted_width = min(max_length + 3, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    except Exception as e:
        st.error(f"Excel oluÅŸturma hatasÄ±: {e}")
        return None

# ==============================================================================
# BUÄDAY GÄ°RÄ°Å ARÅÄ°VÄ° (YENÄ° DÃœZEN)
# ==============================================================================
def show_bugday_giris_arsivi():
    """
    BuÄŸday GiriÅŸ ArÅŸivi
    - Ä°stenen BaÅŸlÄ±k SÄ±ralamasÄ±
    - ID NumaralandÄ±rma
    - HaÅŸere Durumu (Notlardan)
    - GruplandÄ±rÄ±lmÄ±ÅŸ Excel
    - Admin KorumalÄ± DÃ¼zenleme
    """
    st.header("ğŸ—„ï¸ BuÄŸday GiriÅŸ ArÅŸivi & YÃ¶netimi")
    
    df = get_bugday_arsiv()
    
    if df.empty:
        st.info("ğŸ“­ HenÃ¼z arÅŸiv kaydÄ± bulunmuyor.")
        return
    
    # --- FÄ°LTRELEME ---
    with st.expander("ğŸ” KayÄ±t Arama ve Filtreleme", expanded=False):
        col_f1, col_f2 = st.columns(2)
        with col_f1:
            arama = st.text_input("Lot No, Plaka veya TedarikÃ§i Ara", placeholder="Ã–rn: BUGDAY-24...")
        with col_f2:
            silo_filter = st.selectbox("Silo Filtresi", ["TÃ¼mÃ¼"] + list(df['silo_isim'].unique()) if 'silo_isim' in df.columns else ["TÃ¼mÃ¼"])

    # Filtre Uygula
    df_filtered = df.copy()
    if arama:
        df_filtered = df_filtered[
            df_filtered.astype(str).apply(lambda x: x.str.contains(arama, case=False)).any(axis=1)
        ]
    if silo_filter != "TÃ¼mÃ¼":
        df_filtered = df_filtered[df_filtered['silo_isim'] == silo_filter]

    # --- VERÄ° HAZIRLIÄI (EKRAN Ä°Ã‡Ä°N) ---
    # 1. ID SÄ±ralamasÄ± Ekle (1, 2, 3...)
    df_filtered.reset_index(drop=True, inplace=True)
    df_filtered.insert(0, 'ID No', range(1, len(df_filtered) + 1))
    
    # 2. HaÅŸere Durumu (NotlarÄ±n iÃ§inde 'haÅŸere', 'bÃ¶cek' varsa VAR de)
    if 'notlar' in df_filtered.columns:
        df_filtered['HaÅŸere'] = df_filtered['notlar'].apply(
            lambda x: "VAR" if isinstance(x, str) and any(k in x.upper() for k in ["HAÅ", "BÃ–C", "BIT", "CANLI"]) else "TEMÄ°Z"
        )
    else:
        df_filtered['HaÅŸere'] = "-"

    # 3. SayÄ±sal Yuvarlama
    numeric_cols = ['hektolitre', 'protein', 'rutubet', 'gluten', 'gluten_index', 'sedim', 'gecikmeli_sedim', 'sune', 'kirik_ciliz', 'yabanci_tane', 'tonaj', 'fiyat']
    for col in numeric_cols:
        if col in df_filtered.columns:
            df_filtered[col] = pd.to_numeric(df_filtered[col], errors='coerce').round(1)

    # 4. SÃ¼tun AdlandÄ±rma (Mapping) - Senin istediÄŸin baÅŸlÄ±klar
    col_map = {
        'tarih': 'Tarih',
        'lot_no': 'Lot No',
        'plaka': 'Plaka',
        'bugday_cinsi': 'BuÄŸday Cinsi',
        'tedarikci': 'TedarikÃ§i AdÄ±',
        'yore': 'YÃ¶re/BÃ¶lge',
        'silo_isim': 'DÃ¶kÃ¼ldÃ¼ÄŸÃ¼ Silo',
        'tonaj': 'Tonaj',
        'fiyat': 'Fiyat',
        'hektolitre': 'Hektolitre',
        'protein': 'Protein',
        'rutubet': 'Rutubet',
        'gluten': 'Gluten',
        'gluten_index': 'Gluten Ä°ndeks',
        'sedim': 'Sedim',
        'gecikmeli_sedim': 'G. Sedim',
        'sune': 'SÃ¼ne',
        'kirik_ciliz': 'KÄ±rÄ±k & CÄ±lÄ±z',
        'yabanci_tane': 'YabancÄ± Tane',
        'notlar': 'Not'
    }
    
    df_display = df_filtered.rename(columns=col_map)
    
    # Ä°stenen sÃ¼tun sÄ±rasÄ±nÄ± oluÅŸtur
    desired_order = [
        'ID No', 'Tarih', 'Lot No', 'Plaka', 'BuÄŸday Cinsi', 'TedarikÃ§i AdÄ±', 
        'YÃ¶re/BÃ¶lge', 'DÃ¶kÃ¼ldÃ¼ÄŸÃ¼ Silo', 'Tonaj', 'Fiyat',
        'Hektolitre', 'Protein', 'Rutubet', 'Gluten', 'Gluten Ä°ndeks',
        'Sedim', 'G. Sedim', 'SÃ¼ne', 'KÄ±rÄ±k & CÄ±lÄ±z', 'YabancÄ± Tane', 'HaÅŸere', 'Not'
    ]
    
    # Sadece mevcut olanlarÄ± seÃ§ (Hata vermemesi iÃ§in)
    final_cols = [c for c in desired_order if c in df_display.columns]
    df_display = df_display[final_cols]

    # --- TABLO GÃ–STERÄ°MÄ° ---
    st.dataframe(
        df_display,
        use_container_width=True,
        hide_index=True,
        column_config={
            "Tarih": st.column_config.DateColumn("Tarih", format="DD.MM.YYYY"),
            "Tonaj": st.column_config.NumberColumn("Tonaj", format="%.1f"),
            "Fiyat": st.column_config.NumberColumn("Fiyat", format="%.2f â‚º"),
            "ID No": st.column_config.NumberColumn("ID", width="small"),
        }
    )
    
    # --- YENÄ° EXCEL BUTONU ---
    excel_data = export_bugday_giris_ozel_excel(df_filtered) # Orijinal filtrelenmiÅŸ veriyi gÃ¶nderiyoruz
    if excel_data:
        st.download_button(
            label="ğŸ“¥Excel Raporu Ä°ndir",
            data=excel_data,
            file_name=f"Bugday_Giris_Raporu_{datetime.now().strftime('%Y%m%d')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary"
        )
    
    st.divider()

    # ==============================================================================
    # ğŸ”’ YÃ–NETÄ°CÄ° PANELÄ° (ADMIN KÄ°LÄ°DÄ°)
    # ==============================================================================
    if st.session_state.get('user_role') != 'admin':
        return 

    st.subheader("ğŸ› ï¸ KayÄ±t Ä°ÅŸlemleri (YÃ¶netici Paneli)")
    
    # 1. KayÄ±t SeÃ§imi
    lot_list = df_filtered['lot_no'].tolist() if 'lot_no' in df_filtered.columns else []
    
    if not lot_list:
        st.warning("Ä°ÅŸlem yapÄ±lacak kayÄ±t bulunamadÄ±.")
        return

    selected_lot = st.selectbox("Ä°ÅŸlem Yapmak Ä°stediÄŸiniz KaydÄ± SeÃ§iniz (Lot No):", lot_list, key="selected_lot_manage")
    record = df[df['lot_no'] == selected_lot].iloc[0]
    
    df_silo_data = get_silo_data()
    silo_listesi = df_silo_data['isim'].tolist() if not df_silo_data.empty else []
    
    # A) FULL GÃœNCELLEME MODU
    with st.container(border=True):
        st.markdown(f"**ğŸ“ KayÄ±t DÃ¼zenle:** `{selected_lot}`")
        
        with st.form(key="full_update_form"):
            # SATIR 1
            c1, c2, c3, c4 = st.columns(4)
            curr_silo = str(record.get('silo_isim', ''))
            s_idx = silo_listesi.index(curr_silo) if curr_silo in silo_listesi else 0
            
            new_silo = c1.selectbox("DÃ¶kÃ¼ldÃ¼ÄŸÃ¼ Silo", options=silo_listesi, index=s_idx)
            new_cins = c2.text_input("BuÄŸday Cinsi", value=str(record.get('bugday_cinsi', '')))
            new_tonaj = c3.number_input("Tonaj", value=float(record.get('tonaj', 0)), step=0.1)
            new_fiyat = c4.number_input("Fiyat (TL)", value=float(record.get('fiyat', 0)), step=0.1)

            # SATIR 2
            c5, c6, c7, c8 = st.columns(4)
            new_tedarikci = c5.text_input("TedarikÃ§i AdÄ±", value=str(record.get('tedarikci', '')))
            new_plaka = c6.text_input("Plaka", value=str(record.get('plaka', '')))
            new_yore = c7.text_input("YÃ¶re/BÃ¶lge", value=str(record.get('yore', '')))
            new_tarih = c8.text_input("Tarih (YYYY-AA-GG)", value=str(record.get('tarih', '')).split(' ')[0])

            st.markdown("---")
            st.markdown("**ğŸ§ª Fiziksel ve Kimyasal Analizler**")

            # SATIR 3
            l1, l2, l3, l4 = st.columns(4)
            new_hl = l1.number_input("Hektolitre", value=float(record.get('hektolitre', 0)), step=0.1)
            new_protein = l2.number_input("Protein", value=float(record.get('protein', 0)), step=0.1)
            new_rutubet = l3.number_input("Rutubet", value=float(record.get('rutubet', 0)), step=0.1)
            new_gluten = l4.number_input("Gluten", value=float(record.get('gluten', 0)), step=0.1)

            # SATIR 4
            l5, l6, l7, l8 = st.columns(4)
            new_gindex = l5.number_input("Gluten Index", value=float(record.get('gluten_index', 0)), step=1.0)
            new_sedim = l6.number_input("Sedim", value=float(record.get('sedim', 0)), step=1.0)
            new_gsedim = l7.number_input("G. Sedim", value=float(record.get('gecikmeli_sedim', 0)), step=1.0)
            new_sune = l8.number_input("SÃ¼ne", value=float(record.get('sune', 0)), step=0.1)

            # SATIR 5
            l9, l10, l11 = st.columns(3)
            new_kirik = l9.number_input("KÄ±rÄ±k & CÄ±lÄ±z", value=float(record.get('kirik_ciliz', 0)), step=0.1)
            new_yabanci = l10.number_input("YabancÄ± Tane", value=float(record.get('yabanci_tane', 0)), step=0.1)
            new_notlar = l11.text_input("Not", value=str(record.get('notlar', '')))
            
            if st.form_submit_button("âœ… GÃœNCELLE", type="primary"):
                update_payload = {
                    'silo_isim': new_silo, 'bugday_cinsi': new_cins, 'tonaj': new_tonaj,
                    'fiyat': new_fiyat, 'tedarikci': new_tedarikci, 'plaka': new_plaka,
                    'yore': new_yore, 'tarih': new_tarih, 'protein': new_protein,
                    'gluten': new_gluten, 'rutubet': new_rutubet, 'hektolitre': new_hl,
                    'sedim': new_sedim, 'gecikmeli_sedim': new_gsedim, 'gluten_index': new_gindex,
                    'sune': new_sune, 'kirik_ciliz': new_kirik, 'yabanci_tane': new_yabanci,
                    'notlar': new_notlar
                }
                
                success, msg = update_intake_record(selected_lot, update_payload)
                if success:
                    st.success(msg)
                    time.sleep(1.5)
                    st.rerun()
                else:
                    st.error(msg)

    # B) SÄ°LME MODU
    with st.expander("ğŸ—‘ï¸ KaydÄ± Sil", expanded=False):
        st.warning(f"âš ï¸ DÄ°KKAT: `{selected_lot}` numaralÄ± kaydÄ± silmek Ã¼zeresiniz!")
        if st.checkbox("Riskleri anladÄ±m, silmek istiyorum.", key="risk_onayi_box"):
            if st.button("ğŸ”¥ KAYDI KALICI OLARAK SÄ°L", type="primary"):
                success, msg = delete_intake_record(selected_lot)
                if success:
                    st.success(msg)
                    time.sleep(1.5)
                    st.rerun()
                else:
                    st.error(msg)
def show_bugday_spec_yonetimi():
    """BuÄŸday Spesifikasyon YÃ¶netimi - GELÄ°ÅTÄ°RÄ°LMÄ°Å TASARIM"""
    st.header("ğŸ“ BuÄŸday Kalite StandartlarÄ±")
    
    tab1, tab2 = st.tabs(["â• Yeni Standart Ekle", "ğŸ“‹ Mevcut Standartlar"])
    
    with tab1:
        st.subheader("Yeni Standart TanÄ±mla")
        
        # Parametre mapping (ikon + TÃ¼rkÃ§e)
        PARAMETRE_MAP = {
            "protein": {"label": "ğŸ§¬ Protein", "birim": "%"},
            "gluten": {"label": "ğŸŒ¾ Gluten", "birim": "%"},
            "rutubet": {"label": "ğŸ’§ Rutubet", "birim": "%"},
            "hektolitre": {"label": "ğŸ“Š Hektolitre", "birim": "kg/hl"},
            "sedim": {"label": "ğŸ”¬ Sedimantasyon", "birim": "ml"},
            "gluten_index": {"label": "âš—ï¸ Gluten Index", "birim": "%"},
            "sune": {"label": "ğŸ› SÃ¼ne", "birim": "%"},
            "kirik_ciliz": {"label": "ğŸ’” KÄ±rÄ±k & CÄ±lÄ±z", "birim": "%"},
            "yabanci_tane": {"label": "ğŸŒ¿ YabancÄ± Tane", "birim": "%"}
        }
        
        col1, col2 = st.columns(2)
        with col1:
            cins = st.text_input("**ğŸ·ï¸ BuÄŸday Cinsi** *", placeholder="Ã–rn: Bezostaya-1")
        
        with col2:
            param_labels = [f"{v['label']}" for k, v in PARAMETRE_MAP.items()]
            param_keys = list(PARAMETRE_MAP.keys())
            selected_label = st.selectbox("**ğŸ”¬ Kalite Parametresi** *", param_labels)
            param = param_keys[param_labels.index(selected_label)]
            birim = PARAMETRE_MAP[param]['birim']
        
        # DeÄŸer giriÅŸleri - KART TASARIMI
        st.markdown("#### ğŸ“ Standart DeÄŸerler")
        with st.container(border=True):
            col3, col4, col5 = st.columns(3)
            min_val = col3.number_input(f"**Minimum** ({birim})", 0.0, format="%.2f", help="Kabul edilebilir en dÃ¼ÅŸÃ¼k deÄŸer")
            max_val = col4.number_input(f"**Maximum** ({birim})", 0.0, format="%.2f", help="Kabul edilebilir en yÃ¼ksek deÄŸer")
            hedef_val = col5.number_input(f"**Hedef** ({birim})", 0.0, format="%.2f", help="Ä°deal hedef deÄŸer")
        
        st.divider()
        if st.button("ğŸ’¾ Standart Kaydet", type="primary", use_container_width=True):
            if cins and param:
                if save_bugday_spec(cins, param, min_val, max_val, hedef_val):
                    st.success("âœ… Standart kaydedildi!")
                    time.sleep(1)
                    st.rerun()
            else:
                st.error("LÃ¼tfen tÃ¼m zorunlu alanlarÄ± doldurun")
    
    with tab2:
        df_specs = get_all_bugday_specs_dataframe()
        
        if not df_specs.empty:
            # Cinslere gÃ¶re grupla
            cinsler = df_specs['bugday_cinsi'].unique()
            
            PARAMETRE_MAP = {
                "protein": {"label": "ğŸ§¬ Protein", "birim": "%"},
                "gluten": {"label": "ğŸŒ¾ Gluten", "birim": "%"},
                "rutubet": {"label": "ğŸ’§ Rutubet", "birim": "%"},
                "hektolitre": {"label": "ğŸ“Š Hektolitre", "birim": "kg/hl"},
                "sedim": {"label": "ğŸ”¬ Sedimantasyon", "birim": "ml"},
                "gluten_index": {"label": "âš—ï¸ Gluten Index", "birim": "%"},
                "sune": {"label": "ğŸ› SÃ¼ne", "birim": "%"},
                "kirik_ciliz": {"label": "ğŸ’” KÄ±rÄ±k & CÄ±lÄ±z", "birim": "%"},
                "yabanci_tane": {"label": "ğŸŒ¿ YabancÄ± Tane", "birim": "%"}
            }
            
            for cins in cinsler:
                with st.expander(f"ğŸŒ¾ **{cins}**", expanded=False):
                    cins_df = df_specs[df_specs['bugday_cinsi'] == cins].copy()
                    
                    # Parametreleri TÃ¼rkÃ§e etiketle
                    cins_df['Parametre'] = cins_df['parametre'].apply(
                        lambda x: PARAMETRE_MAP.get(x, {"label": x})['label']
                    )
                    
                    # GÃ¶sterim iÃ§in yeniden dÃ¼zenle
                    display_df = cins_df[['Parametre', 'min_deger', 'max_deger', 'hedef_deger']].copy()
                    display_df.columns = ['Parametre', 'Min', 'Max', 'Hedef']
                    
                    st.dataframe(
                        display_df, 
                        use_container_width=True, 
                        hide_index=True,
                        column_config={
                            "Parametre": st.column_config.TextColumn("Parametre", width="medium"),
                            "Min": st.column_config.NumberColumn("Min", format="%.2f"),
                            "Max": st.column_config.NumberColumn("Max", format="%.2f"),
                            "Hedef": st.column_config.NumberColumn("Hedef â­", format="%.2f")
                        }
                    )
                    
                    # Silme butonu - ONAY Ä°LE
                    col_a, col_b = st.columns([3, 1])
                    with col_b:
                        if st.button(f"ğŸ—‘ï¸ Sil", key=f"del_{cins}", type="secondary", use_container_width=True):
                            if f"confirm_delete_{cins}" not in st.session_state:
                                st.session_state[f"confirm_delete_{cins}"] = True
                                st.warning(f"âš ï¸ '{cins}' standardÄ±nÄ± silmek istediÄŸinize emin misiniz?")
                                st.rerun()
                    
                    # Onay mesajÄ± gÃ¶sterildiyse
                    if st.session_state.get(f"confirm_delete_{cins}", False):
                        col_x, col_y = st.columns(2)
                        with col_x:
                            if st.button("âœ… Evet, Sil", key=f"confirm_yes_{cins}", type="primary"):
                                if delete_bugday_spec_group(cins):
                                    st.success(f"âœ… {cins} silindi")
                                    del st.session_state[f"confirm_delete_{cins}"]
                                    time.sleep(1)
                                    st.rerun()
                        with col_y:
                            if st.button("âŒ Ä°ptal", key=f"confirm_no_{cins}"):
                                del st.session_state[f"confirm_delete_{cins}"]
                                st.rerun()
        else:
            st.info("ğŸ“­ HenÃ¼z standart tanÄ±mlanmamÄ±ÅŸ")
            st.markdown("""
            **ğŸ’¡ Ä°pucu:** Yeni bir standart eklemek iÃ§in yukarÄ±daki **'Yeni Standart Ekle'** sekmesini kullanÄ±n.
            """)
# --------------------------------------------------------------------------
# BUÄDAY YÃ–NETÄ°M MERKEZÄ° (YENÄ° EKLENEN ANA FONKSÄ°YON)
# --------------------------------------------------------------------------
def show_wheat_yonetimi():
    """
    BuÄŸday Operasyon Merkezi
    TÃ¼m giriÅŸ, analiz, paÃ§al ve stok sÃ¼reÃ§lerinin yÃ¶netildiÄŸi ana ekran.
    """
    
    # 1. BaÅŸlÄ±k AlanÄ± (YeÅŸil/TarÄ±m TemasÄ±)
    st.markdown("""
    <div style='background-color: #E8F5E9; padding: 15px; border-radius: 10px; margin-bottom: 20px; border-left: 5px solid #2E7D32;'>
        <h2 style='color: #1B5E20; margin:0;'>ğŸŒ¾ BuÄŸday Operasyon Merkezi</h2>
        <p style='color: #4CAF50; margin:0; font-size: 14px;'>Hammadde GiriÅŸ, Kalite YÃ¶netimi, PaÃ§al ve Stok Takibi</p>
    </div>
    """, unsafe_allow_html=True)

    # 2. Yatay MenÃ¼ (Senin belirlediÄŸin yapÄ±)
    secim = st.radio(
        "ModÃ¼l SeÃ§iniz:",
        [
            "ğŸš› GiriÅŸ & Kalite OperasyonlarÄ±", 
            "âš—ï¸ PaÃ§al (Blend) YÃ¶netimi", 
            "ğŸ“¤ Stok Ã‡Ä±kÄ±ÅŸÄ±", 
            "ğŸ“‚ Veri TabanÄ± & Ä°zlenebilirlik"
        ],
        horizontal=True,
        label_visibility="collapsed"
    )
    
    st.markdown("---")

    # 3. YÃ¶nlendirmeler ve Sekmeler
    
    # --- A) GÄ°RÄ°Å & KALÄ°TE ---
    if secim == "ğŸš› GiriÅŸ & Kalite OperasyonlarÄ±":
        # Ä°Ã§ Sekmeler
        tab1, tab2, tab3 = st.tabs(["ğŸ“ Spek & Hedefler", "ğŸ“¥ Hammadde GiriÅŸ", "ğŸ§ª TavlÄ± Analiz GiriÅŸi"])
        
        with tab1:
            with st.container(border=True):
                show_bugday_spec_yonetimi()
        
        with tab2:
            with st.container(border=True):
                show_mal_kabul()
                
        with tab3:
            with st.container(border=True):
                show_tavli_analiz()

    # --- B) PAÃ‡AL (BLEND) YÃ–NETÄ°MÄ° ---
    elif secim == "âš—ï¸ PaÃ§al (Blend) YÃ¶netimi":
        try:
            import app.modules.calculations as calculations
            
            tab_p1, tab_p2 = st.tabs(["ğŸ§® PaÃ§al HesaplayÄ±cÄ±", "ğŸ“œ PaÃ§al GeÃ§miÅŸi"])
            
            with tab_p1:
                with st.container(border=True):
                    if hasattr(calculations, 'show_pacal_hesaplayici'):
                        calculations.show_pacal_hesaplayici()
                    else:
                        st.warning("âš ï¸ PaÃ§al HesaplayÄ±cÄ± modÃ¼lÃ¼ bulunamadÄ±.")
            
            with tab_p2:
                with st.container(border=True):
                    if hasattr(calculations, 'show_pacal_gecmisi'):
                        calculations.show_pacal_gecmisi()
                    else:
                        st.warning("âš ï¸ PaÃ§al GeÃ§miÅŸi modÃ¼lÃ¼ bulunamadÄ±.")
                        
        except ImportError:
            st.error("âš ï¸ 'app.modules.calculations' modÃ¼lÃ¼ yÃ¼klenemedi!")
        except Exception as e:
            st.error(f"âš ï¸ Bir hata oluÅŸtu: {e}")

    # --- C) STOK Ã‡IKIÅI ---
    elif secim == "ğŸ“¤ Stok Ã‡Ä±kÄ±ÅŸÄ±":
        with st.container(border=True):
            show_stok_cikis()

    # --- D) VERÄ° TABANI & Ä°ZLENEBÄ°LÄ°RLÄ°K ---
    elif secim == "ğŸ“‚ Veri TabanÄ± & Ä°zlenebilirlik":
        tab_db1, tab_db2 = st.tabs(["ğŸ“’ GiriÅŸ ArÅŸivi", "ğŸ”„ Stok Hareketleri"])
        
        with tab_db1:
            with st.container(border=True):
                show_bugday_giris_arsivi()
                
        with tab_db2:
            with st.container(border=True):
                show_stok_hareketleri()
# ==============================================================================
# Ã–ZEL EXCEL RAPOR MOTORU (TAVLI BUÄDAY - ONDALIK DÃœZELTÄ°LMÄ°Å & BAÅLIKLAR DÃœZELTÄ°LDÄ°)
# ==============================================================================
def export_tavli_ozel_excel(df):
    """
    TavlÄ± analizler iÃ§in Ã¶zel gruplandÄ±rÄ±lmÄ±ÅŸ baÅŸlÄ±klÄ± Excel Ã¼retir.
    DÃœZELTME: BaÅŸ harfler bÃ¼yÃ¼tÃ¼ldÃ¼ (Protein, Gluten, Rutubet, Sedim)
    """
    try:
        from io import BytesIO
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = "TavlÄ± Analiz Raporu"

        # --- TASARIM TANIMLARI ---
        structure = [
            {
                "group": "TEMEL BÄ°LGÄ°LER",
                "color": "4472C4", # Mavi
                "cols": [
                    ("Tarih", "tarih"),
                    ("Silo", "silo_isim"),
                    ("Tonaj", "analiz_tonaj"),
                    ("Notlar", "notlar")
                ]
            },
            {
                "group": "KÄ°MYASAL ANALÄ°ZLER",
                "color": "ED7D31", # Turuncu
                "cols": [
                    ("Protein", "protein"),      # DÃœZELTÄ°LDÄ°
                    ("Gluten", "gluten"),        # DÃœZELTÄ°LDÄ°
                    ("Rutubet", "rutubet"),      # DÃœZELTÄ°LDÄ°
                    ("G. Ä°ndeks", "gluten_index"), 
                    ("Sedim", "sedim"),          # DÃœZELTÄ°LDÄ°
                    ("G. Sedim", "g_sedim"),
                    ("FN", "fn"),
                    ("FFN", "ffn"),
                    ("Amilograf", "amilograph")
                ]
            },
            {
                "group": "FARINOGRAPH ANALÄ°ZLERÄ°",
                "color": "70AD47", # YeÅŸil
                "cols": [
                    ("Su Kal. (F)", "su_kaldirma_f"),
                    ("GeliÅŸme", "gelisme_suresi"),
                    ("Stabilite", "stabilite"),
                    ("YumuÅŸama", "yumusama")
                ]
            },
            {
                "group": "EXTENSOGRAPH ANALÄ°ZLERÄ°",
                "color": "A5A5A5", # Gri
                "cols": [
                    ("Su Kal. (E)", "su_kaldirma_e"),
                    # 45 DK
                    ("Enerji (45)", "enerji45"), ("DirenÃ§ (45)", "direnc45"), ("Taban (45)", "taban45"),
                    # 90 DK
                    ("Enerji (90)", "enerji90"), ("DirenÃ§ (90)", "direnc90"), ("Taban (90)", "taban90"),
                    # 135 DK
                    ("Enerji (135)", "enerji135"), ("DirenÃ§ (135)", "direnc135"), ("Taban (135)", "taban135")
                ]
            }
        ]

        # --- STÄ°LLER ---
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        header_font = Font(bold=True, color="FFFFFF", size=11)
        sub_header_font = Font(bold=True, color="000000", size=10)
        
        # --- BAÅLIKLARI YAZMA ---
        current_col = 1
        for group in structure:
            start_col = current_col
            num_cols = len(group["cols"])
            end_col = start_col + num_cols - 1
            
            # 1. Ãœst BaÅŸlÄ±k (Merge)
            ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
            cell = ws.cell(row=1, column=start_col, value=group["group"])
            cell.fill = PatternFill("solid", fgColor=group["color"])
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
            
            for c in range(start_col, end_col + 1):
                ws.cell(row=1, column=c).border = thin_border

            # 2. Alt BaÅŸlÄ±klar
            for i, (col_name, db_key) in enumerate(group["cols"]):
                cell_sub = ws.cell(row=2, column=start_col + i, value=col_name)
                cell_sub.font = sub_header_font
                cell_sub.alignment = Alignment(horizontal="center", vertical="center")
                cell_sub.border = thin_border
                cell_sub.fill = PatternFill("solid", fgColor="E7E6E6")

            current_col += num_cols

        # --- VERÄ°LERÄ° YAZMA ---
        for r_idx, row_data in enumerate(df.to_dict('records'), start=3):
            current_col = 1
            for group in structure:
                for col_name, db_key in group["cols"]:
                    val = row_data.get(db_key, "")
                    
                    # Tarih formatÄ±
                    if db_key == "tarih" and val:
                        try: val = pd.to_datetime(val).strftime('%d.%m.%Y %H:%M')
                        except: pass
                    
                    # SayÄ±sal yuvarlama (12.168 -> 12.2)
                    try:
                        if db_key not in ["tarih", "silo_isim", "notlar"] and val is not None and val != "":
                            val = float(val)
                            val = round(val, 1) # Excel'e temiz gitmesi iÃ§in yuvarla
                    except: pass

                    cell = ws.cell(row=r_idx, column=current_col, value=val)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal="center")
                    current_col += 1

        # --- SÃœTUN GENÄ°ÅLÄ°KLERÄ° ---
        for i, col in enumerate(ws.columns, 1):
            max_length = 0
            column_letter = get_column_letter(i)
            
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except: pass
            
            adjusted_width = min(max_length + 3, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    except Exception as e:
        st.error(f"Excel oluÅŸturma hatasÄ±: {e}")
        return None

# ==============================================================================
# TAVLI ANALÄ°Z ARÅÄ°VÄ° (BAÅLIKLAR DÃœZELTÄ°LDÄ°: Protein, Gluten, Rutubet...)
# ==============================================================================
def show_tavli_analiz_arsivi():
    """
    TavlÄ± BuÄŸday Analiz GeÃ§miÅŸi
    - TÃ¼rkÃ§e ve BÃ¼yÃ¼k BaÅŸ Harfli BaÅŸlÄ±klar (Protein, Gluten...)
    - SayÄ±sal Yuvarlama (12.2)
    - Admin Yetkili DÃ¼zenleme
    """
    st.markdown("### ğŸ§ª TavlÄ± BuÄŸday Analiz GeÃ§miÅŸi")
    
    # Veriyi Ã‡ek
    df = get_tavli_analizler()
    
    if df.empty:
        st.info("ğŸ“­ HenÃ¼z tavlÄ± analiz kaydÄ± bulunmuyor.")
        return

    # --- FÄ°LTRELEME ---
    with st.expander("ğŸ” Filtreleme SeÃ§enekleri", expanded=False):
        c1, c2 = st.columns(2)
        with c1:
            silo_filter = st.selectbox("Silo Filtrele", ["TÃ¼mÃ¼"] + list(df['silo_isim'].unique()))
        with c2:
            pass

    df_show = df.copy()
    if silo_filter != "TÃ¼mÃ¼":
        df_show = df_show[df_show['silo_isim'] == silo_filter]

    # --- VERÄ° HAZIRLIÄI (YUVARLAMA VE BAÅLIKLAR) ---
    # 1. TÃ¼m sayÄ±sal kolonlarÄ± 1 ondalÄ±ÄŸa yuvarla
    numeric_cols = [
        'protein', 'gluten', 'rutubet', 'gluten_index', 'sedim', 'g_sedim', 'fn', 'ffn', 'amilograph',
        'su_kaldirma_f', 'gelisme_suresi', 'stabilite', 'yumusama', 'su_kaldirma_e',
        'enerji45', 'direnc45', 'taban45', 'enerji90', 'direnc90', 'taban90', 'enerji135', 'direnc135', 'taban135',
        'analiz_tonaj'
    ]
    
    for col in numeric_cols:
        if col in df_show.columns:
            # Ã–nce sayÄ±ya Ã§evir (hata varsa NaN yap), sonra yuvarla
            df_show[col] = pd.to_numeric(df_show[col], errors='coerce').round(1)

    # 2. Tablo GÃ¶sterimi Ä°Ã§in Kopya Al ve BaÅŸlÄ±klarÄ± TÃ¼rkÃ§eleÅŸtir/DÃ¼zelt
    df_display = df_show.copy()
    
    # Ä°ÅTE BURASI: Ekrandaki Tablo BaÅŸlÄ±klarÄ±nÄ± DÃ¼zeltiyoruz
    col_map = {
        'silo_isim': 'Silo',
        'analiz_tonaj': 'Tonaj',
        'tarih': 'Tarih',
        'notlar': 'Notlar',
        # Kimyasal (Ä°stediÄŸin BÃ¼yÃ¼k BaÅŸ Harfler)
        'protein': 'Protein',   # DÃœZELTÄ°LDÄ°
        'gluten': 'Gluten',     # DÃœZELTÄ°LDÄ°
        'rutubet': 'Rutubet',   # DÃœZELTÄ°LDÄ°
        'sedim': 'Sedim',       # DÃœZELTÄ°LDÄ°
        'gluten_index': 'G. Ä°ndeks',
        'g_sedim': 'G. Sedim',
        'amilograph': 'Amilograf',
        'fn': 'FN',
        'ffn': 'FFN',
        # Farino
        'su_kaldirma_f': 'Su Kal. (F)',
        'gelisme_suresi': 'GeliÅŸme',
        'stabilite': 'Stabilite',
        'yumusama': 'YumuÅŸama',
        # Extenso
        'su_kaldirma_e': 'Su Kal. (E)',
        'enerji45': 'Enerji (45)', 'direnc45': 'DirenÃ§ (45)', 'taban45': 'Taban (45)',
        'enerji90': 'Enerji (90)', 'direnc90': 'DirenÃ§ (90)', 'taban90': 'Taban (90)',
        'enerji135': 'Enerji (135)', 'direnc135': 'DirenÃ§ (135)', 'taban135': 'Taban (135)'
    }
    
    # Sadece var olan sÃ¼tunlarÄ± yeniden adlandÄ±r
    df_display = df_display.rename(columns=col_map)

    # --- EKRAN TABLOSU ---
    # Not: column_config kullanarak formatÄ± zorluyoruz (%.1f)
    st.dataframe(
        df_display,
        use_container_width=True,
        hide_index=True,
        column_config={
            "Tarih": st.column_config.DatetimeColumn("Tarih", format="DD.MM.YYYY HH:mm"),
            "Tonaj": st.column_config.NumberColumn("Tonaj", format="%.1f"),
            "Protein": st.column_config.NumberColumn("Protein", format="%.1f"),  # DÃœZELTÄ°LDÄ°
            "Gluten": st.column_config.NumberColumn("Gluten", format="%.1f"),    # DÃœZELTÄ°LDÄ°
            "Rutubet": st.column_config.NumberColumn("Rutubet", format="%.1f"),  # DÃœZELTÄ°LDÄ°
            "Sedim": st.column_config.NumberColumn("Sedim", format="%.1f"),      # DÃœZELTÄ°LDÄ°
            "G. Ä°ndeks": st.column_config.NumberColumn("G. Ä°ndeks", format="%.1f"),
            "Su Kal. (F)": st.column_config.NumberColumn("Su Kal. (F)", format="%.1f"),
        }
    )
    
    # --- Ã–ZEL EXCEL BUTONU ---
    excel_data = export_tavli_ozel_excel(df_show)
    if excel_data:
        st.download_button(
            label="ğŸ“¥Excel Raporu Ä°ndir",
            data=excel_data,
            file_name=f"Tavli_Analiz_Raporu_{datetime.now().strftime('%Y%m%d')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary"
        )

    st.divider()

    # ==========================================================================
    # ğŸ”’ YÃ–NETÄ°CÄ° PANELÄ° (ADMIN ONLY)
    # ==========================================================================
    if st.session_state.get('user_role') != 'admin':
        return

    st.subheader("ğŸ› ï¸ KayÄ±t DÃ¼zenleme (Admin Paneli)")
    
    # KayÄ±t SeÃ§imi
    record_list = df_show.to_dict('records')
    def format_func(row):
        t_str = pd.to_datetime(row['tarih']).strftime('%d.%m %H:%M') if pd.notnull(row['tarih']) else "Tarih Yok"
        return f"{row.get('silo_isim','?')} - {t_str} ({row.get('analiz_tonaj',0)} Ton)"

    selected_record = st.selectbox("DÃ¼zenlenecek KaydÄ± SeÃ§in:", record_list, format_func=format_func)
    
    if selected_record:
        df_silo_data = get_silo_data()
        silo_opts = df_silo_data['isim'].tolist() if not df_silo_data.empty else []
        
        with st.form(key="tavli_edit_form"):
            st.markdown(f"**DÃ¼zenlenen KayÄ±t:** `{format_func(selected_record)}`")
            
            # --- BÃ–LÃœM 1: TEMEL BÄ°LGÄ°LER ---
            st.markdown("#### 1. Temel Bilgiler")
            col_t1, col_t2, col_t3, col_t4 = st.columns(4)
            
            curr_silo = selected_record.get('silo_isim')
            s_idx = silo_opts.index(curr_silo) if curr_silo in silo_opts else 0
            
            new_silo = col_t1.selectbox("Silo (DÄ°KKAT!)", options=silo_opts, index=s_idx)
            new_tonaj = col_t2.number_input("Tonaj (DÄ°KKAT!)", value=float(selected_record.get('analiz_tonaj', 0)), step=0.1)
            new_tarih = col_t3.text_input("Tarih", value=str(selected_record.get('tarih')))
            new_not = col_t4.text_input("Notlar", value=str(selected_record.get('notlar', '')))

            st.markdown("---")
            
            # --- BÃ–LÃœM 2: DETAYLI ANALÄ°ZLER ---
            tab_kimya, tab_farino, tab_extenso = st.tabs(["ğŸ§ª Kimyasal", "ğŸ“ˆ Farinograph", "ğŸ“Š Extensograph"])
            
            def get_val(k, default=0.0): return float(selected_record.get(k, default))

            with tab_kimya:
                k1, k2, k3, k4 = st.columns(4)
                n_protein = k1.number_input("Protein", value=get_val('protein'), step=0.1)
                n_gluten = k2.number_input("Gluten", value=get_val('gluten'), step=0.1)
                n_rutubet = k3.number_input("Rutubet", value=get_val('rutubet'), step=0.1)
                n_sedim = k4.number_input("Sedim", value=get_val('sedim'), step=1.0)
                
                k5, k6, k7, k8 = st.columns(4)
                n_gsedim = k5.number_input("G. Sedim", value=get_val('g_sedim'), step=1.0)
                n_gindex = k6.number_input("G. Index", value=get_val('gluten_index'), step=1.0)
                n_fn = k7.number_input("FN", value=get_val('fn'), step=1.0)
                n_ffn = k8.number_input("FFN", value=get_val('ffn'), step=1.0)
                n_amilo = st.number_input("Amilograph", value=get_val('amilograph'), step=10.0)

            with tab_farino:
                f1, f2, f3, f4 = st.columns(4)
                n_su_kaldirma_f = f1.number_input("Su KaldÄ±rma (F)", value=get_val('su_kaldirma_f'), step=0.1)
                n_gelisme = f2.number_input("GeliÅŸme", value=get_val('gelisme_suresi'), step=0.1)
                n_stabilite = f3.number_input("Stabilite", value=get_val('stabilite'), step=0.1)
                n_yumusama = f4.number_input("YumuÅŸama", value=get_val('yumusama'), step=1.0)

            with tab_extenso:
                st.write("**Extensograph Verileri**")
                n_su_kaldirma_e = st.number_input("Su KaldÄ±rma (E)", value=get_val('su_kaldirma_e'), step=0.1)
                
                with st.expander("45. Dakika", expanded=False):
                    ex1, ex2, ex3 = st.columns(3)
                    n_e45 = ex1.number_input("Enerji (45)", value=get_val('enerji45'), key="ne45")
                    n_d45 = ex2.number_input("DirenÃ§ (45)", value=get_val('direnc45'), key="nd45")
                    n_t45 = ex3.number_input("Taban (45)", value=get_val('taban45'), key="nt45")
                
                with st.expander("90. Dakika", expanded=False):
                    ex4, ex5, ex6 = st.columns(3)
                    n_e90 = ex4.number_input("Enerji (90)", value=get_val('enerji90'), key="ne90")
                    n_d90 = ex5.number_input("DirenÃ§ (90)", value=get_val('direnc90'), key="nd90")
                    n_t90 = ex6.number_input("Taban (90)", value=get_val('taban90'), key="nt90")
                    
                with st.expander("135. Dakika", expanded=False):
                    ex7, ex8, ex9 = st.columns(3)
                    n_e135 = ex7.number_input("Enerji (135)", value=get_val('enerji135'), key="ne135")
                    n_d135 = ex8.number_input("DirenÃ§ (135)", value=get_val('direnc135'), key="nd135")
                    n_t135 = ex9.number_input("Taban (135)", value=get_val('taban135'), key="nt135")

            c_btn1, c_btn2 = st.columns([1, 4])
            with c_btn1:
                submit_update = st.form_submit_button("âœ… GÃœNCELLE", type="primary")
            
            if submit_update:
                new_data = {
                    'silo_isim': new_silo, 'analiz_tonaj': new_tonaj, 'tarih': new_tarih, 'notlar': new_not,
                    'protein': n_protein, 'gluten': n_gluten, 'rutubet': n_rutubet, 'sedim': n_sedim,
                    'g_sedim': n_gsedim, 'gluten_index': n_gindex, 'fn': n_fn, 'ffn': n_ffn, 'amilograph': n_amilo,
                    'su_kaldirma_f': n_su_kaldirma_f, 'gelisme_suresi': n_gelisme, 'stabilite': n_stabilite, 'yumusama': n_yumusama,
                    'su_kaldirma_e': n_su_kaldirma_e,
                    'enerji45': n_e45, 'direnc45': n_d45, 'taban45': n_t45,
                    'enerji90': n_e90, 'direnc90': n_d90, 'taban90': n_t90,
                    'enerji135': n_e135, 'direnc135': n_d135, 'taban135': n_t135
                }
                
                success, msg = update_tavli_record_backend(selected_record, new_data)
                if success:
                    st.success(msg)
                    time.sleep(1.5)
                    st.rerun()
                else:
                    st.error(msg)
        
        with st.expander("ğŸ—‘ï¸ Bu KaydÄ± Sil", expanded=False):
            st.warning(f"Bu iÅŸlem **{selected_record['silo_isim']}** silosundan **{selected_record['analiz_tonaj']}** tonluk stoÄŸu dÃ¼ÅŸecektir.")
            if st.button("ğŸ”¥ KALICI OLARAK SÄ°L"):
                success, msg = delete_tavli_record_backend(selected_record)
                if success:
                    st.success(msg)
                    time.sleep(1.5)
                    st.rerun()
                else:
                    st.error(msg)













































