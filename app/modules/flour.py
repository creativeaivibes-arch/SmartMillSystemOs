import streamlit as st
import pandas as pd
import time
from datetime import datetime
import plotly.express as px
import plotly.graph_objects as go

from app.core.database import fetch_data, add_data, get_conn
from app.core.utils import turkce_karakter_duzelt
from app.core.config import INPUT_LIMITS, TERMS, get_limit

try:
    from app.modules.reports import create_un_maliyet_pdf_report, download_styled_excel
    
except ImportError:
    def create_un_maliyet_pdf_report(*args): return None
    def download_styled_excel(*args): pass

# --- AYARLAR (CONFIG) - MAGIC NUMBERS ---
FLOUR_CONFIG = {
    'SPEC_ACTIVE_STATE': 1,       # VarsayÄ±lan aktiflik durumu
    'DEFAULT_TABLE_HEIGHT': 400,  # Tablo yÃ¼kseklikleri
    'DECIMAL_PRECISION': 2,       # VarsayÄ±lan ondalÄ±k hassasiyet
    'DEFAULT_ANALYSIS_COUNT': 10, # VarsayÄ±lan gÃ¶sterilecek analiz sayÄ±sÄ±
    'DATE_FORMAT_DB': '%Y-%m-%d %H:%M:%S',
    'DATE_FORMAT_DISPLAY': '%d.%m.%Y %H:%M'
}


# --- YENÄ° EKLENEN HELPER ---
def get_active_production_lots():
    """
    Sevkiyat iÃ§in referans alÄ±nacak ÃœRETÄ°M (PRD) kayÄ±tlarÄ±nÄ± Ã§eker.
    KaynaÄŸÄ± 'un_analiz' tablosudur (Laboratuvar Ãœretim KayÄ±tlarÄ±).
    """
    try:
        # 1. un_analiz tablosunu Ã§ek (Force refresh ile en gÃ¼ncel hali)
        df = fetch_data("un_analiz", force_refresh=True)
        if df.empty: 
            return []
        
        # 2. Sadece 'ÃœRETÄ°M' tipindeki kayÄ±tlarÄ± filtrele
        # (Ã‡Ã¼nkÃ¼ Sevkiyat, daha Ã¶nce Ã¼retilmiÅŸ ve analizi yapÄ±lmÄ±ÅŸ mallardan yapÄ±lÄ±r)
        if 'islem_tipi' in df.columns:
            df = df[df['islem_tipi'] == "ÃœRETÄ°M"]
        
        if df.empty:
            return []

        # 3. Tarihe gÃ¶re sÄ±rala (En yeni en Ã¼stte)
        if 'tarih' in df.columns:
            df['tarih'] = pd.to_datetime(df['tarih'], errors='coerce')
            df = df.sort_values('tarih', ascending=False)
            
        lot_list = []
        for _, row in df.iterrows():
            try:
                # Lot No (PRD-...)
                lot = str(row.get('lot_no', ''))
                if not lot or lot.lower() == 'nan': continue

                # ÃœrÃ¼n AdÄ± / Marka
                # Hem 'un_markasi' hem 'un_cinsi_marka' alanlarÄ±na bakÄ±yoruz
                marka = row.get('un_markasi', '')
                if not marka: 
                    marka = row.get('un_cinsi_marka', '-')
                
                # Tarih
                tarih_str = row['tarih'].strftime('%d.%m %H:%M') if pd.notnull(row['tarih']) else "-"
                
                # Format: PRD-260210... | PÄ±rlanta | 10.02 14:30
                label = f"{lot} | {marka} | {tarih_str}"
                lot_list.append(label)
            except:
                continue
            
        return lot_list
    except Exception as e:
        return []
def get_active_mixing_batches():
    """
    PaÃ§al (Mixing) kayÄ±tlarÄ±nÄ± Ã§eker.
    Ãœretim analizi girerken 'Hangi PaÃ§al KullanÄ±ldÄ±?' sorusu iÃ§in.
    """
    try:
        df = fetch_data("mixing_batches")
        if df.empty: 
            return []
        
        if 'tarih' in df.columns:
            df['tarih'] = pd.to_datetime(df['tarih'])
            df = df.sort_values('tarih', ascending=False)
            
        batch_list = []
        for _, row in df.iterrows():
            # Format: ReÃ§ete AdÄ± | Tarih | MIX-ID
            tarih_str = row['tarih'].strftime('%d.%m %H:%M') if pd.notnull(row['tarih']) else "-"
            # Batch ID ve ÃœrÃ¼n AdÄ±nÄ± birleÅŸtir
            label = f"{row.get('urun_adi', 'PaÃ§al')} | {tarih_str} | {row.get('batch_id', '?')}"
            batch_list.append(label)
            
        return batch_list
    except Exception as e:
        return []

def get_un_maliyet_gecmisi():
    """Maliyet geÃ§miÅŸini dÃ¶ndÃ¼r"""
    df = fetch_data("un_maliyet_hesaplamalari")
    if df.empty:
        return pd.DataFrame()
    if 'tarih' in df.columns:
        df['tarih'] = pd.to_datetime(df['tarih'], errors='coerce')
        df = df.sort_values('tarih', ascending=False)
    return df

# --- VERÄ°TABANI Ä°ÅLEMLERÄ° (DRY PRENSÄ°BÄ°NE UYGUN) ---
def _update_spec_table(df_new):
    """YardÄ±mcÄ± Fonksiyon: Spec tablosunu gÃ¼venli ÅŸekilde gÃ¼nceller"""
    try:
        conn = get_conn()
        conn.update(worksheet="un_spekleri", data=df_new)
        return True
    except Exception as e:
        st.error(f"VeritabanÄ± GÃ¼ncelleme HatasÄ±: {e}")
        return False

def save_spec(un_cinsi, parametre, min_val, max_val, hedef_val, tolerans):
    """Spec ekleme veya gÃ¼ncelleme iÅŸlemini tek merkezden yÃ¶netir"""
    try:
        df = fetch_data("un_spekleri")
        
        # Yeni kayÄ±t verisi
        new_row = {
            'un_cinsi': un_cinsi, 
            'parametre': parametre, 
            'min_deger': float(min_val), 
            'max_deger': float(max_val), 
            'hedef_deger': float(hedef_val), 
            'tolerans': float(tolerans), 
            'aktif': FLOUR_CONFIG['SPEC_ACTIVE_STATE']
        }
        
        if df.empty:
            # Tablo boÅŸsa direkt ekle (add_data kullanabiliriz ama update standardÄ± iÃ§in DataFrame oluÅŸturuyoruz)
            df_new = pd.DataFrame([new_row])
            return _update_spec_table(df_new)
        
        # Mevcut kaydÄ± ara
        mask = (df['un_cinsi'] == un_cinsi) & (df['parametre'] == parametre)
        
        if mask.any():
            # Varsa GÃœNCELLE
            df.loc[mask, ['min_deger', 'max_deger', 'hedef_deger', 'tolerans', 'aktif']] = \
                [float(min_val), float(max_val), float(hedef_val), float(tolerans), FLOUR_CONFIG['SPEC_ACTIVE_STATE']]
        else:
            # Yoksa EKLE (DataFrame'e append et)
            new_df_row = pd.DataFrame([new_row])
            df = pd.concat([df, new_df_row], ignore_index=True)
            
        return _update_spec_table(df)

    except Exception as e:
        st.error(f"KayÄ±t Ä°ÅŸlemi HatasÄ±: {e}")
        return False

def delete_spec_group(un_cinsi):
    """Belirtilen un cinsine ait tÃ¼m spekleri siler"""
    try:
        df = fetch_data("un_spekleri")
        if df.empty: return True
        
        # Filtreleme mantÄ±ÄŸÄ± ile silme (O un cinsi OLMAYANLARI tut)
        df_new = df[df['un_cinsi'] != un_cinsi]
        
        # EÄŸer satÄ±r sayÄ±sÄ± deÄŸiÅŸtiyse gÃ¼ncelle
        if len(df_new) < len(df):
            return _update_spec_table(df_new)
        return True
        
    except Exception as e:
        st.error(f"Silme HatasÄ±: {e}")
        return False

def get_all_specs_dataframe():
    """TÃ¼m spekleri listelemek iÃ§in veriyi Ã§eker ve formatlar"""
    df = fetch_data("un_spekleri")
    if df.empty: return pd.DataFrame()
    
    # SÃ¼tun isimlerini kullanÄ±cÄ± dostu hale getir
    return df.rename(columns={
        'un_cinsi': 'Un Cinsi', 
        'parametre': 'Parametre',
        'min_deger': 'Min', 
        'hedef_deger': 'Hedef', 
        'max_deger': 'Max'
    })

def show_spec_yonetimi():
    """Un Kalite SpesifikasyonlarÄ± (Spec) EkranÄ± - GÃ¼venli ve Validasyonlu"""
    st.markdown("### ğŸ¯ Un Kalite SpesifikasyonlarÄ± (Spec)")
    
    # --- 1. GÃœVENLÄ° VERÄ° Ã‡EKME ---
    df_spek = pd.DataFrame()
    try:
        raw_data = fetch_data("un_spekleri")
        if isinstance(raw_data, pd.DataFrame):
            df_spek = raw_data
    except Exception as e:
        st.warning(f"Veri baÄŸlantÄ± hatasÄ±: {e}")

    # --- 2. LÄ°STE HAZIRLIÄI ---
    un_listesi = set()
    if not df_spek.empty and 'un_cinsi' in df_spek.columns:
        try:
            items = df_spek['un_cinsi'].dropna().unique().tolist()
            un_listesi.update(items)
        except: pass
    
    all_types = sorted(list(un_listesi))

    # --- 3. ARAYÃœZ VE GÄ°RÄ°Å KONTROLÃœ (VALIDASYON EKLENDÄ°) ---
    col_sel, col_add = st.columns([2, 1])
    with col_sel:
        secilen_urun = st.selectbox(
            "DÃ¼zenlenecek Un Cinsini SeÃ§iniz", 
            ["(SeÃ§iniz/Yeni Ekle)"] + all_types,
            key="spec_select_box"
        )
        
    if secilen_urun == "(SeÃ§iniz/Yeni Ekle)":
        with col_add:
            ham_isim = st.text_input("â• Yeni Un TanÄ±mla", placeholder="Ã–rn: Tam BuÄŸday").strip()
            
            # [GÃœVENLÄ°K] TÃ¼rkÃ§e karakter dÃ¼zeltme ve standartlaÅŸtÄ±rma
            if ham_isim:
                # Ä°sim temizliÄŸi (Ã–rn: "tam buÄŸday" -> "TAM BUGDAY")
                temiz_isim = turkce_karakter_duzelt(ham_isim).upper()
                
                # [VALIDASYON] Uzunluk ve tekrar kontrolÃ¼
                if len(temiz_isim) < 3:
                    st.caption("âš ï¸ Ä°sim en az 3 karakter olmalÄ±.")
                    secilen_urun = None
                elif temiz_isim in all_types:
                    st.toast("âš ï¸ Bu un cinsi zaten kayÄ±tlÄ±, mevcut kayda yÃ¶nlendirildi.", icon="â„¹ï¸")
                    secilen_urun = temiz_isim # Mevcut olana yÃ¶nlendir
                else:
                    secilen_urun = temiz_isim
            else:
                secilen_urun = None

    # EÄŸer geÃ§erli bir seÃ§im yoksa dur
    if not secilen_urun:
        st.info("ğŸ‘† LÃ¼tfen dÃ¼zenlemek veya oluÅŸturmak iÃ§in bir un cinsi seÃ§in.")
        if not df_spek.empty:
            st.divider()
            st.caption("ğŸ“‹ Sistemde KayÄ±tlÄ± Spekler")
            # Ã–nizleme tablosu
            st.dataframe(
                df_spek[['un_cinsi', 'parametre', 'hedef_deger']].head(10), 
                use_container_width=True,
                hide_index=True
            )
        return

    # --- 4. DÃœZENLEME FORMU ---
    st.divider()
    
    # SeÃ§ilen Ã¼rÃ¼nÃ¼n mevcut deÄŸerlerini Ã§ek
    current_specs = {}
    if not df_spek.empty and 'un_cinsi' in df_spek.columns:
        df_filtered = df_spek[df_spek['un_cinsi'] == secilen_urun]
        for _, row in df_filtered.iterrows():
            current_specs[row['parametre']] = row

    param_groups = {
        "Kimyasal Analizler": [
            ("protein", "Protein (%)"), ("rutubet", "Rutubet (%)"), ("kul", "KÃ¼l (%)"),
            ("gluten", "Gluten (%)"), ("gluten_index", "Gluten Index"), ("sedim", "Sedim (ml)"),
            ("gecikmeli_sedim", "Gecikmeli Sedim (ml)"), ("fn", "DÃ¼ÅŸme SayÄ±sÄ± (FN)"),
            ("ffn", "F.F.N"), ("nisasta_zedelenmesi", "NiÅŸasta Zedelenmesi")
        ],
        "Farinograph & Amilograph": [
            ("su_kaldirma_f", "Su KaldÄ±rma (Farino) (%)"), ("gelisme_suresi", "GeliÅŸme SÃ¼resi (dk)"),
            ("stabilite", "Stabilite (dk)"), ("yumusama", "YumuÅŸama Derecesi (FU)"),
            ("amilograph", "Amilograph (AU)")
        ],
        "Extensograph": [
            ("enerji45", "Enerji (45 dk)"), ("direnc45", "DirenÃ§ (45 dk)"), ("taban45", "Uzama/Taban (45 dk)"),
            ("enerji90", "Enerji (90 dk)"), ("direnc90", "DirenÃ§ (90 dk)"), ("taban90", "Uzama/Taban (90 dk)"),
            ("enerji135", "Enerji (135 dk)"), ("direnc135", "DirenÃ§ (135 dk)"), ("taban135", "Uzama/Taban (135 dk)"),
            ("su_kaldirma_e", "Su KaldÄ±rma (Extenso) (%)")
        ]
    }

    st.markdown(f"### ğŸ› ï¸ DÃ¼zenleme: **{secilen_urun}**")
    
    with st.form("spec_editor_comprehensive"):
        tabs = st.tabs(list(param_groups.keys()))
        input_keys = []
        
        # Helper: GÃ¼venli float Ã§eviri
        def safe_float(val):
            try: return float(val)
            except: return 0.0

        for idx, (group_name, params) in enumerate(param_groups.items()):
            with tabs[idx]:
                for p_key, p_label in params:
                    cur = current_specs.get(p_key, {})
                    val_min = safe_float(cur.get('min_deger', 0.0))
                    val_tgt = safe_float(cur.get('hedef_deger', 0.0))
                    val_max = safe_float(cur.get('max_deger', 0.0))
                    
                    c1, c2, c3, c4 = st.columns([2, 1, 1, 1])
                    with c1: st.markdown(f"**{p_label}**")
                    # Config'den gelen hassasiyet kullanÄ±labilir veya standart 2 hane
                    with c2: st.number_input("Min", value=val_min, key=f"min_{p_key}", step=0.1, format="%.2f", label_visibility="collapsed")
                    with c3: st.number_input("Hedef", value=val_tgt, key=f"tgt_{p_key}", step=0.1, format="%.2f", label_visibility="collapsed")
                    with c4: st.number_input("Max", value=val_max, key=f"max_{p_key}", step=0.1, format="%.2f", label_visibility="collapsed")
                    input_keys.append(p_key)
        
        st.divider()
        if st.form_submit_button("ğŸ’¾ Kaydet / GÃ¼ncelle", type="primary", use_container_width=True):
            saved_count = 0
            # Progress bar ile kullanÄ±cÄ±ya geri bildirim
            prog_bar = st.progress(0)
            
            for i, p_key in enumerate(input_keys):
                s_min = st.session_state.get(f"min_{p_key}", 0.0)
                s_tgt = st.session_state.get(f"tgt_{p_key}", 0.0)
                s_max = st.session_state.get(f"max_{p_key}", 0.0)
                
                # Sadece deÄŸer girilmiÅŸse kaydet (0,0,0 olanlarÄ± pas geÃ§erek veritabanÄ±nÄ± ÅŸiÅŸirme)
                if s_min > 0 or s_tgt > 0 or s_max > 0:
                    if save_spec(secilen_urun, p_key, s_min, s_max, s_tgt, 0):
                        saved_count += 1
                
                # Ä°lerleme Ã§ubuÄŸunu gÃ¼ncelle
                prog_bar.progress((i + 1) / len(input_keys))
            
            prog_bar.empty()
            
            if saved_count > 0:
                st.success(f"âœ… **{secilen_urun}** iÃ§in {saved_count} parametre baÅŸarÄ±yla gÃ¼ncellendi.")
                time.sleep(1)
                st.rerun()
            else:
                st.warning("âš ï¸ DeÄŸiÅŸiklik algÄ±lanmadÄ± veya tÃ¼m deÄŸerler 0 girildi.")

    # Silme Butonu (Sadece Admin)
    if st.session_state.get("user_role") == "admin":
        st.divider()
        with st.expander("ğŸ—‘ï¸ Tehlikeli BÃ¶lge"):
            if st.button("Bu ÃœrÃ¼n TanÄ±mÄ±nÄ± ve TÃ¼m Speklerini Sil", key="del_spec_main", type="primary"):
                if delete_spec_group(secilen_urun):
                    st.success("TanÄ±m Silindi!")
                    time.sleep(1)
                    st.rerun()

def export_un_analiz_ozel_excel(df):
    """
    Un Analiz ArÅŸivi iÃ§in Ã¶zel gruplandÄ±rÄ±lmÄ±ÅŸ Excel Ã¼retir.
    YapÄ±: [SEVKÄ°YAT/TAKÄ°P] + [NUMUNE BÄ°LGÄ°LERÄ°] + [KÄ°MYASAL] + [FARINO] + [EXTENSO]
    """
    try:
        from io import BytesIO
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = "Un Analiz ve Sevkiyat"

        # --- TASARIM TANIMLARI ---
        structure = [
            {
                "group": "Ä°ZLENEBÄ°LÄ°RLÄ°K & SEVKÄ°YAT",  # <-- YENÄ° GRUP
                "color": "7030A0", # Mor
                "cols": [
                    ("ID NO", "id_counter"),
                    ("TARÄ°H", "tarih"),
                    ("Ä°ÅLEM", "islem_tipi"),
                    ("MÃœÅTERÄ°", "musteri_adi"),
                    ("PLAKA/ÅOFÃ–R", "plaka_no"),
                    ("KAYNAK PARTÄ°", "kaynak_parti_no")
                ]
            },
            {
                "group": "NUMUNE DETAYLARI",
                "color": "4472C4", # Mavi
                "cols": [
                    ("LOT NO", "lot_no"),
                    ("UN CÄ°NSÄ°", "un_cinsi_marka"),
                    ("MARKA", "un_markasi"),
                    ("SÄ°LO", "uretim_silosu"),
                    ("NOTLAR", "notlar")
                ]
            },
            {
                "group": "KÄ°MYASAL ANALÄ°ZLER",
                "color": "ED7D31", # Turuncu
                "cols": [
                    ("Protein", "protein"),
                    ("Rutubet", "rutubet"),
                    ("Gluten", "gluten"),
                    ("Gluten Index", "gluten_index"),
                    ("Sedim", "sedim"),
                    ("G.Sedim", "gecikmeli_sedim"),
                    ("F.N", "fn"),
                    ("F.F.N", "ffn"),
                    ("Amilograph", "amilograph"),
                    ("KÃ¼l", "kul"),
                    ("NiÅŸasta Zed.", "nisasta_zedelenmesi")
                ]
            },
            {
                "group": "FARINOGRAPH",
                "color": "70AD47", # YeÅŸil
                "cols": [
                    ("Su KaldÄ±rma", "su_kaldirma_f"),
                    ("GeliÅŸme SÃ¼resi", "gelisme_suresi"),
                    ("Stabilite", "stabilite"),
                    ("YumuÅŸama", "yumusama")
                ]
            },
            # Extenso aynÄ± kalÄ±yor...
             {
                "group": "EXTENSOGRAPH",
                "color": "A5A5A5", # Gri
                "cols": [
                    ("Su KaldÄ±rma (E)", "su_kaldirma_e"),
                    ("DirenÃ§ (45)", "direnc45"), ("Taban (45)", "taban45"), ("Enerji (45)", "enerji45"),
                    ("DirenÃ§ (90)", "direnc90"), ("Taban (90)", "taban90"), ("Enerji (90)", "enerji90"),
                    ("DirenÃ§ (135)", "direnc135"), ("Taban (135)", "taban135"), ("Enerji (135)", "enerji135")
                ]
            }
        ]

        # --- STÄ°LLER ---
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        header_font = Font(bold=True, color="FFFFFF", size=11)
        sub_header_font = Font(bold=True, color="000000", size=10)
        
        # --- BAÅLIKLARI YAZMA ---
        current_col = 1
        for group in structure:
            start_col = current_col
            num_cols = len(group["cols"])
            end_col = start_col + num_cols - 1
            
            ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
            cell = ws.cell(row=1, column=start_col, value=group["group"])
            cell.fill = PatternFill("solid", fgColor=group["color"])
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
            
            for c in range(start_col, end_col + 1):
                ws.cell(row=1, column=c).border = thin_border

            for i, (col_name, db_key) in enumerate(group["cols"]):
                cell_sub = ws.cell(row=2, column=start_col + i, value=col_name)
                cell_sub.font = sub_header_font
                cell_sub.alignment = Alignment(horizontal="center", vertical="center")
                cell_sub.border = thin_border
                cell_sub.fill = PatternFill("solid", fgColor="E7E6E6")

            current_col += num_cols

        # --- VERÄ°LERÄ° YAZMA ---
        records = df.to_dict('records')
        for r_idx, row_data in enumerate(records, start=3):
            current_col = 1
            for group in structure:
                for col_name, db_key in group["cols"]:
                    
                    if db_key == "id_counter":
                        val = r_idx - 2
                    else:
                        val = row_data.get(db_key, "")
                    
                    if db_key == "tarih" and val:
                        try: val = pd.to_datetime(val).strftime('%d.%m.%Y %H:%M')
                        except: pass
                    
                    # SayÄ±sal yuvarlama (MÃ¼ÅŸteri adÄ± gibi text alanlarÄ±nÄ± bozmadan)
                    try:
                        if isinstance(val, (int, float)):
                            val = round(float(val), 2)
                        elif val and db_key not in ["tarih", "lot_no", "islem_tipi", "uretim_silosu", "notlar", "musteri_adi", "plaka_no", "kaynak_parti_no", "un_cinsi_marka", "un_markasi"]:
                            # Sadece analiz deÄŸerlerini yuvarlamaya Ã§alÄ±ÅŸ
                            try: val = round(float(val), 2)
                            except: pass
                    except: pass

                    cell = ws.cell(row=r_idx, column=current_col, value=val)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal="center")
                    current_col += 1

        for i, col in enumerate(ws.columns, 1):
            column_letter = get_column_letter(i)
            ws.column_dimensions[column_letter].width = 15

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    except Exception as e:
        st.error(f"Excel oluÅŸturma hatasÄ±: {e}")
        return None
    
    

def save_un_analiz(lot_no, islem_tipi, **analiz_degerleri):
    try:
        df_check = fetch_data("un_analiz")
        if not df_check.empty and 'lot_no' in df_check.columns:
            if lot_no in df_check['lot_no'].values:
                return False, f"Bu lot numarasÄ± zaten kayÄ±tlÄ±: {lot_no}"
        data = {
            'lot_no': str(lot_no),
            'islem_tipi': islem_tipi,
            'tarih': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            **analiz_degerleri
        }
        if add_data("un_analiz", data):
            return True, "KayÄ±t BaÅŸarÄ±lÄ±"
        return False, "KayÄ±t BaÅŸarÄ±sÄ±z"
    except Exception as e:
        return False, f"Hata: {str(e)}"
def update_un_analiz_record(old_lot_no, new_data):
    """Un analiz kaydÄ±nÄ± gÃ¼nceller"""
    try:
        conn = get_conn()
        df = fetch_data("un_analiz")
        
        # Lot numarasÄ±na gÃ¶re satÄ±rÄ± bul
        if not df.empty and 'lot_no' in df.columns:
            # Pandas indexini bul
            idx_list = df.index[df['lot_no'].astype(str) == str(old_lot_no)].tolist()
            
            if idx_list:
                idx = idx_list[0]
                # Verileri gÃ¼ncelle
                for key, val in new_data.items():
                    df.at[idx, key] = val
                
                conn.update(worksheet="un_analiz", data=df)
                return True, "âœ… KayÄ±t baÅŸarÄ±yla gÃ¼ncellendi."
            else:
                return False, "KayÄ±t bulunamadÄ±."
        return False, "VeritabanÄ± boÅŸ."
    except Exception as e:
        return False, f"GÃ¼ncelleme HatasÄ±: {str(e)}"

def delete_un_analiz_record(lot_no):
    """Un analiz kaydÄ±nÄ± siler"""
    try:
        conn = get_conn()
        df = fetch_data("un_analiz")
        
        if not df.empty and 'lot_no' in df.columns:
            # O lot numarasÄ± dÄ±ÅŸÄ±ndakileri al (Filtreleme ile silme)
            df_new = df[df['lot_no'].astype(str) != str(lot_no)]
            conn.update(worksheet="un_analiz", data=df_new)
            return True, "ğŸ—‘ï¸ KayÄ±t silindi."
        return False, "VeritabanÄ± hatasÄ±."
    except Exception as e:
        return False, f"Silme HatasÄ±: {str(e)}"

def show_un_analiz_kaydi():
    if st.session_state.get('user_role') not in ["admin", "operations", "quality"]:
        st.warning("â›” Yetkisiz EriÅŸim")
        return
    
    st.header("ğŸ“ Un Analiz & Sevkiyat KaydÄ±")
    
    # --- 1. Ä°ÅLEM TÄ°PÄ° VE AKILLI LOT ---
    islem_tipi = st.selectbox("Ä°ÅŸlem Tipi SeÃ§iniz *", ["ÃœRETÄ°M", "SEVKÄ°YAT", "NUMUNE", "ÅÄ°KAYET", "Ä°ADE"])
    
    prefix_map = {
        "ÃœRETÄ°M": "PRD",     # Production Analysis
        "SEVKÄ°YAT": "SHIP",  # Shipment
        "NUMUNE": "SAMPLE",  # Random Sample
        "Ä°ADE": "RTRN",      # Return
        "ÅÄ°KAYET": "CLAIM"   # Claim
    }
    current_prefix = prefix_map.get(islem_tipi, "UN")
    timestamp_str = datetime.now().strftime('%y%m%d%H%M')
    auto_lot = f"{current_prefix}-{timestamp_str}"

    col1, col2 = st.columns([1, 1], gap="medium")
    
    # --- SOL KOLON: KÄ°MLÄ°K BÄ°LGÄ°LERÄ° ---
    with col1:
        st.subheader("ğŸ“‹ KayÄ±t KÃ¼nyesi")
        
        st.info(f"**Otomatik Lot:** `{auto_lot}`")
        lot_no = st.text_input("Lot NumarasÄ± *", value=auto_lot)
        analiz_tarihi = st.date_input("Analiz Tarihi", datetime.now())
        
        # --- DÄ°NAMÄ°K ALANLAR ---
        kaynak_parti = None
        musteri_adi = None
        plaka_no = None
        uretim_silosu = None
        
        # A) SEVKÄ°YAT MODU
        if islem_tipi == "SEVKÄ°YAT":
            st.markdown("ğŸšš **Sevkiyat DetaylarÄ±**")
            
            musteri_adi = st.text_input("MÃ¼ÅŸteri / Firma AdÄ± *")
            plaka_no = st.text_input("AraÃ§ PlakasÄ± / ÅofÃ¶r *")
            
            # Kaynak: ÃœRETÄ°M (PRD) KayÄ±tlarÄ±
            prod_lots = get_active_production_lots()
            secilen_kaynak = st.selectbox("Hangi Ãœretimden Sevk Ediliyor? (PRD Referans)", ["(Stoktan / KarÄ±ÅŸÄ±k)"] + prod_lots)
            
            # PRD ID'sini ayÄ±kla (String parse: "PRD-123 | Ekmeklik..." -> "PRD-123")
            if secilen_kaynak != "(Stoktan / KarÄ±ÅŸÄ±k)":
                try: kaynak_parti = secilen_kaynak.split(' | ')[0].strip()
                except: kaynak_parti = secilen_kaynak

        # B) ÃœRETÄ°M MODU (DÃœZELTÄ°LDÄ°: ARTIK PAÃ‡AL SORUYOR)
        elif islem_tipi == "ÃœRETÄ°M":
            st.markdown("ğŸ­ **Ãœretim KaynaÄŸÄ±**")
            
            # Kaynak: PAÃ‡AL (MIX) KayÄ±tlarÄ±
            # BURASI DÃœZELDÄ°: ArtÄ±k PaÃ§al listesini getiriyor
            mix_batches = get_active_mixing_batches()
            secilen_parti = st.selectbox("ÃœRETÄ°M PAÃ‡ALI SEÃ‡Ä°NÄ°Z (MIX Referans)", ["(BaÄŸÄ±msÄ±z / HazÄ±r Stok)"] + mix_batches)
            
            # MIX ID'sini ayÄ±kla
            if secilen_parti != "(BaÄŸÄ±msÄ±z / HazÄ±r Stok)":
                try: 
                    # "isim | tarih | ID" formatÄ± varsayÄ±lÄ±yor, son parÃ§a ID
                    kaynak_parti = secilen_parti.split(' | ')[-1].strip()
                except: 
                    kaynak_parti = secilen_parti

            # Silo
            df_silo = fetch_data("silolar") 
            if not df_silo.empty:
                col_name = 'isim' if 'isim' in df_silo.columns else df_silo.columns[0]
                silo_list = ["(BelirtilmemiÅŸ)"] + df_silo[col_name].tolist()
                uretim_silosu = st.selectbox("Ãœretim Silosu", silo_list)
            else:
                uretim_silosu = st.text_input("Ãœretim Silosu", placeholder="Silo No")

        st.divider()
        
        # UN CÄ°NSÄ° / MARKA (SADELEÅTÄ°RÄ°LDÄ°)
        st.markdown("ğŸŒ¾ **ÃœrÃ¼n TanÄ±mÄ±**")
        # Spec seÃ§imi kaldÄ±rÄ±ldÄ±, sadece Marka/Ticari Ä°sim kaldÄ±
        un_markasi = st.text_input("Un MarkasÄ± (Ticari) *", placeholder="Ã–rn: PÄ±rlanta, YÄ±ldÄ±z, LÃ¼ks Ekmeklik")
        
        # VeritabanÄ±nda 'un_cinsi_marka' zorunlu olduÄŸu iÃ§in, girilen markayÄ± oraya da yazÄ±yoruz
        un_cinsi_marka = un_markasi 
            
        notlar = st.text_area("Notlar")

    # --- SAÄ KOLON: LABORATUVAR ANALÄ°ZLERÄ° ---
    with col2:
        st.subheader("ğŸ§ª Laboratuvar Analizleri")
        
        # --- TAB DÃœZENÄ° ---
        tab1, tab2, tab3 = st.tabs(["ğŸ§ª KÄ°MYASAL (TÃ¼mÃ¼)", "ğŸ“ˆ FARINOGRAPH", "ğŸ“Š EXTENSOGRAPH"])

        # TAB 1: KÄ°MYASAL
        with tab1:
            c1, c2, c3 = st.columns(3)
            with c1:
                protein = st.number_input("Protein (%)", 0.0, 20.0, 0.0, 0.1) # VarsayÄ±lan 0
                rutubet = st.number_input("Rutubet (%)", 0.0, 20.0, 0.0, 0.1)
                gluten = st.number_input("Gluten (%)", 0.0, 50.0, 0.0, 0.1)
                gluten_index = st.number_input("Gluten Index", 0.0, 100.0, 0.0, 1.0)
            with c2:
                sedim = st.number_input("Sedim (ml)", 0.0, 100.0, 0.0, 1.0)
                g_sedim = st.number_input("Gecikmeli Sedim", 0.0, 100.0, 0.0, 1.0)
                fn = st.number_input("DÃ¼ÅŸme SayÄ±sÄ± (FN)", 0.0, 999.0, 0.0, 1.0)
                ffn = st.number_input("F.F.N", 0.0, 999.0, 0.0, 1.0)
            with c3:
                amilo = st.number_input("Amilograph (AU)", 0.0, value=0.0)
                nisasta = st.number_input("NiÅŸasta Zed.", 0.0, value=0.0)
                kul = st.number_input("KÃ¼l (%)", 0.0, value=0.0, step=0.001, format="%.3f")

        # TAB 2: FARINOGRAPH
        with tab2:
            c1, c2 = st.columns(2)
            with c1:
                f_su = st.number_input("Su KaldÄ±rma (%)", 0.0, value=0.0)
                f_gelisme = st.number_input("GeliÅŸme SÃ¼resi (dk)", 0.0, value=0.0)
            with c2:
                f_stab = st.number_input("Stabilite (dk)", 0.0, value=0.0)
                f_yumus = st.number_input("YumuÅŸama (FU)", 0.0, value=0.0)

        # TAB 3: EXTENSOGRAPH
        with tab3:
            su_e = st.number_input("Su KaldÄ±rma (Extenso) (%)", value=0.0)
            st.divider()
            
            t1, t2, t3 = st.columns(3)
            with t1:
                st.markdown("**45. Dakika**")
                e45_d = t1.number_input("DirenÃ§ (45)", value=0.0)
                e45_t = t1.number_input("Taban (45)", value=0.0)
                e45_e = t1.number_input("Enerji (45)", value=0.0)
            with t2:
                st.markdown("**90. Dakika**")
                e90_d = t2.number_input("DirenÃ§ (90)", value=0.0)
                e90_t = t2.number_input("Taban (90)", value=0.0)
                e90_e = t2.number_input("Enerji (90)", value=0.0)
            with t3:
                st.markdown("**135. Dakika**")
                e135_d = t3.number_input("DirenÃ§ (135)", value=0.0)
                e135_t = t3.number_input("Taban (135)", value=0.0)
                e135_e = t3.number_input("Enerji (135)", value=0.0)

    st.divider()
    
    # --- KAYIT BUTONU ---
    btn_text = "ğŸšš SEVKÄ°YATI KAYDET" if islem_tipi == "SEVKÄ°YAT" else "âœ… ANALÄ°ZÄ° KAYDET"
    
    if st.button(btn_text, type="primary", use_container_width=True):
        # 1. ZORUNLU ALANLAR
        if not lot_no:
            st.error("âš ï¸ Lot NumarasÄ± boÅŸ olamaz!")
            return
        
        if not un_markasi:
            st.error("âš ï¸ LÃ¼tfen 'Un MarkasÄ±' giriniz.")
            return
            
        # 2. SEVKÄ°YAT KONTROLÃœ
        if islem_tipi == "SEVKÄ°YAT":
            if not musteri_adi or not plaka_no:
                st.error("âš ï¸ Sevkiyat iÃ§in MÃ¼ÅŸteri AdÄ± ve Plaka zorunludur!")
                return
        
        # 3. ANALÄ°Z KONTROLÃœ (SEVKÄ°YATTA ZORUNLULUK KALKTI AMA DEÄER GÄ°RÄ°LEBÄ°LÄ°R)
        # Sadece Ãœretim ve Numune'de '0' kontrolÃ¼ yapalÄ±m mÄ±?
        # KullanÄ±cÄ± 'bazen' girmek istemeyebilir, o yÃ¼zden 0'a izin veriyoruz ama logluyoruz.
        # EÄŸer kesin zorunluluk istiyorsan burayÄ± aÃ§abiliriz.
        
        # 4. VERÄ° PAKETLEME
        analiz_data = {
            'un_cinsi_marka': un_markasi,  # ArtÄ±k marka adÄ±nÄ± kullanÄ±yor
            'un_markasi': un_markasi, 
            'uretim_silosu': uretim_silosu,
            'kaynak_parti_no': kaynak_parti, # MIX ID veya PRD ID buraya geliyor
            'musteri_adi': musteri_adi,
            'plaka_no': plaka_no,
            
            # Analizler
            'protein': protein, 'rutubet': rutubet, 'gluten': gluten, 'gluten_index': gluten_index,
            'sedim': sedim, 'gecikmeli_sedim': g_sedim, 'fn': fn, 'ffn': ffn,
            'amilograph': amilo, 'nisasta_zedelenmesi': nisasta, 'kul': kul,
            
            # Farino
            'su_kaldirma_f': f_su, 'gelisme_suresi': f_gelisme, 'stabilite': f_stab, 'yumusama': f_yumus,
            
            # Extenso
            'su_kaldirma_e': su_e,
            'direnc45': e45_d, 'taban45': e45_t, 'enerji45': e45_e,
            'direnc90': e90_d, 'taban90': e90_t, 'enerji90': e90_e,
            'direnc135': e135_d, 'taban135': e135_t, 'enerji135': e135_e,
            
            'notlar': notlar
        }
        
        ok, msg = save_un_analiz(lot_no, islem_tipi, **analiz_data)
        if ok:
            st.success(f"âœ… Ä°ÅŸlem BaÅŸarÄ±lÄ±! ({islem_tipi})")
            time.sleep(1)
            st.rerun()
        else:
            st.error(f"âŒ {msg}")

def show_un_analiz_kayitlari():
    """Un Analiz ArÅŸivi - Sevkiyat ve Ä°zlenebilirlik Dahil"""
    st.header("ğŸ“š Un Analiz ve Sevkiyat KayÄ±tlarÄ±")
    
    df = fetch_data("un_analiz")
    if df.empty:
        st.info("ğŸ“­ HenÃ¼z kayÄ±tlÄ± iÅŸlem bulunmamaktadÄ±r.")
        return

    # --- VERÄ° HAZIRLIÄI ---
    if 'tarih' in df.columns:
        df['tarih'] = pd.to_datetime(df['tarih'], errors='coerce')
        df = df.sort_values('tarih', ascending=False)
    
    df.reset_index(drop=True, inplace=True)
    df.insert(0, 'ID NO', range(1, len(df) + 1))

    # SayÄ±sal dÃ¶nÃ¼ÅŸtÃ¼rme (Sadece analiz sÃ¼tunlarÄ± iÃ§in)
    numeric_cols = [
        'protein', 'rutubet', 'gluten', 'gluten_index', 'sedim', 'gecikmeli_sedim',
        'fn', 'ffn', 'amilograph', 'kul', 'nisasta_zedelenmesi',
        'su_kaldirma_f', 'gelisme_suresi', 'stabilite', 'yumusama', 'su_kaldirma_e',
        'direnc45', 'taban45', 'enerji45', 'direnc90', 'taban90', 'enerji90',
        'direnc135', 'taban135', 'enerji135'
    ]
    for col in numeric_cols:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce')

    # BaÅŸlÄ±klarÄ± EÅŸle (YENÄ° SÃœTUNLAR EKLENDÄ°)
    col_map = {
        'tarih': 'TARÄ°H', 'lot_no': 'LOT NO', 'islem_tipi': 'Ä°ÅLEM',
        'un_cinsi_marka': 'UN CÄ°NSÄ°', 'uretim_silosu': 'SÄ°LO', 'notlar': 'NOTLAR',
        # -- YENÄ°LER --
        'musteri_adi': 'MÃœÅTERÄ°',
        'plaka_no': 'PLAKA',
        'kaynak_parti_no': 'KAYNAK (PRD)',
        # -- ANALÄ°ZLER --
        'protein': 'Protein', 'rutubet': 'Rutubet', 'gluten': 'Gluten', 
        'gluten_index': 'Gluten Index', 'sedim': 'Sedim', 'gecikmeli_sedim': 'G.Sedim',
        'fn': 'F.N', 'ffn': 'F.F.N', 'amilograph': 'Amilograph', 'kul': 'KÃ¼l',
        'nisasta_zedelenmesi': 'NiÅŸasta Zed.',
        'su_kaldirma_f': 'Su KaldÄ±rma (F)', 'gelisme_suresi': 'GeliÅŸme SÃ¼resi',
        'stabilite': 'Stabilite', 'yumusama': 'YumuÅŸama Derecesi',
        'su_kaldirma_e': 'Su KaldÄ±rma (E)',
        'direnc45': 'DirenÃ§ (45)', 'taban45': 'Taban (45)', 'enerji45': 'Enerji (45)',
        'direnc90': 'DirenÃ§ (90)', 'taban90': 'Taban (90)', 'enerji90': 'Enerji (90)',
        'direnc135': 'DirenÃ§ (135)', 'taban135': 'Taban (135)', 'enerji135': 'Enerji (135)'
    }
    
    df_display = df.rename(columns=col_map)
    
    # Ä°stenen SÃ¼tun SÄ±ralamasÄ± (YENÄ°LENMÄ°Å)
    desired_cols = [
        'ID NO', 'TARÄ°H', 'Ä°ÅLEM', 'MÃœÅTERÄ°', 'PLAKA', 'KAYNAK (PRD)', # Ã–ne aldÄ±k
        'LOT NO', 'UN CÄ°NSÄ°', 'SÄ°LO', 'NOTLAR',
        'Protein', 'Rutubet', 'Gluten', 'Gluten Index', 'Sedim', 'G.Sedim',
        'F.N', 'F.F.N', 'Amilograph', 'KÃ¼l', 'NiÅŸasta Zed.',
        'Su KaldÄ±rma (F)', 'GeliÅŸme SÃ¼resi', 'Stabilite', 'YumuÅŸama Derecesi',
        'Su KaldÄ±rma (E)',
        'DirenÃ§ (45)', 'Taban (45)', 'Enerji (45)',
        'DirenÃ§ (90)', 'Taban (90)', 'Enerji (90)',
        'DirenÃ§ (135)', 'Taban (135)', 'Enerji (135)'
    ]
    
    # Mevcut olmayan sÃ¼tunlarÄ± atla (Hata almamak iÃ§in)
    final_cols = [c for c in desired_cols if c in df_display.columns]
    df_display = df_display[final_cols]

    st.subheader(f"ğŸ“Š Toplam KayÄ±t: {len(df)}")
    
    # TABLO GÃ–STERÄ°MÄ°
    st.dataframe(
        df_display, 
        use_container_width=True, 
        hide_index=True,
        column_config={
            "TARÄ°H": st.column_config.DatetimeColumn("TARÄ°H", format="DD.MM.YYYY HH:mm"),
            "Protein": st.column_config.NumberColumn("Protein", format="%.2f"),
            "KÃ¼l": st.column_config.NumberColumn("KÃ¼l", format="%.3f"),
            "Gluten": st.column_config.NumberColumn("Gluten", format="%.1f"),
            "Rutubet": st.column_config.NumberColumn("Rutubet", format="%.1f"),
            # Ä°ÅŸlem tipine gÃ¶re renklendirme veya ikon eklenebilir ama basit tutuyoruz
        }
    )
    
    # Excel Butonu
    excel_data = export_un_analiz_ozel_excel(df) 
    if excel_data:
        st.download_button(
            label="ğŸ“¥ Excel Ä°ndir (Sevkiyat DetaylÄ±)",
            data=excel_data,
            file_name=f"SmartMill_Rapor_{datetime.now().strftime('%Y%m%d')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary"
        )
    
    st.divider()

    # (YÃ¶netici Paneli Kodu Eski Halinde Kalabilir - DeÄŸiÅŸiklik yok)
    if st.session_state.get('user_role') != 'admin':
        return

    st.subheader("ğŸ› ï¸ KayÄ±t Ä°ÅŸlemleri (YÃ¶netici Paneli)")
    
    lot_list = df['lot_no'].tolist() if 'lot_no' in df.columns else []
    if not lot_list: 
        st.warning("DÃ¼zenlenecek kayÄ±t bulunamadÄ±.")
        return

    def format_func(lot):
        row = df[df['lot_no'] == lot].iloc[0]
        t_str = pd.to_datetime(row['tarih']).strftime('%d.%m %H:%M') if pd.notnull(row['tarih']) else ""
        return f"{lot} - {row.get('islem_tipi','?')} ({t_str})"

    selected_lot = st.selectbox("DÃ¼zenlenecek KaydÄ± SeÃ§in (Lot No):", lot_list, format_func=format_func)
    
    # ... (Silme butonu mantÄ±ÄŸÄ± aynen devam eder) ...
    # Silme butonunu tekrar yazmÄ±yorum, eski kodundaki "B) SÄ°LME BUTONU" kÄ±smÄ±nÄ± koru.
    with st.expander("ğŸ—‘ï¸ KaydÄ± Sil", expanded=False):
        st.warning(f"âš ï¸ DÄ°KKAT: `{selected_lot}` numaralÄ± kaydÄ± silmek Ã¼zeresiniz!")
        if st.checkbox("Riskleri anladÄ±m, silmek istiyorum.", key="un_del_confirm"):
            if st.button("ğŸ”¥ KAYDI KALICI OLARAK SÄ°L", type="primary"):
                success, msg = delete_un_analiz_record(selected_lot)
                if success:
                    st.success(msg)
                    time.sleep(1.5)
                    st.rerun()
                else:
                    st.error(msg)

def delete_un_maliyet_record(tarih_val):
    """Maliyet kaydÄ±nÄ± tarihe gÃ¶re siler"""
    try:
        conn = get_conn()
        df = fetch_data("un_maliyet_hesaplamalari")
        if df.empty: return False

        # Tarih sÃ¼tununu stringe Ã§evirip karÅŸÄ±laÅŸtÄ±ralÄ±m (EÅŸleÅŸme garantisi iÃ§in)
        df['tarih'] = df['tarih'].astype(str)
        tarih_str = str(tarih_val)
        
        # EÅŸleÅŸmeyenleri tut (Silme mantÄ±ÄŸÄ±)
        df_new = df[df['tarih'] != tarih_str]
        
        # EÄŸer satÄ±r sayÄ±sÄ± azaldÄ±ysa silme baÅŸarÄ±lÄ±dÄ±r
        if len(df_new) < len(df):
            conn.update(worksheet="un_maliyet_hesaplamalari", data=df_new)
            return True
        return False
    except Exception as e:
        return False

def save_un_maliyet(data):
    """Maliyet hesaplamasÄ±nÄ± kaydet"""
    try:
        data['tarih'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        data['kullanici'] = st.session_state.get('username', 'Sistem')
        return add_data("un_maliyet_hesaplamalari", data)
    except: 
        return False


def show_un_maliyet_hesaplama():
    """Un Maliyet Hesaplama - SADECE HESAPLAMA"""
    st.header("ğŸ§® Un Maliyet Hesaplama")
    
    currency = "TL"
    
    col_filter1, col_filter2 = st.columns(2)    
    with col_filter1:
        ay_listesi = ["OCAK", "ÅUBAT", "MART", "NÄ°SAN", "MAYIS", "HAZÄ°RAN", 
                     "TEMMUZ", "AÄUSTOS", "EYLÃœL", "EKÄ°M", "KASIM", "ARALIK"]
        secilen_ay = st.selectbox("Hesaplama AyÄ±", ay_listesi, index=datetime.now().month - 1)
    
    with col_filter2:
        yil_listesi = list(range(2024, 2037))
        secilen_yil = st.selectbox("Hesaplama YÄ±lÄ±", yil_listesi, index=2)
    
    st.divider()
    st.subheader(f"Un Maliyeti Hesapla - {secilen_ay} {secilen_yil}")
    
    col1, col2, col3 = st.columns(3, gap="medium")
    
    with col1:
        st.markdown("#### ğŸ“‹ TEMEL BÄ°LGÄ°LER")
        un_cesidi = st.text_input("Un Ã‡eÅŸidi *", value="Ekmeklik", placeholder="Ã–rn: Ekmeklik, Pizza")
        bugday_maliyet = st.number_input("BuÄŸday PaÃ§al (TL/KG) *", min_value=0.0, value=14.60, step=0.01, format="%.2f")
        aylik_kirilan = st.number_input("AylÄ±k KÄ±rÄ±lan (Ton) *", min_value=0.0, value=3000.0, step=0.1, format="%.1f")
        randiman = st.number_input("RandÄ±man (%) *", min_value=0.0, max_value=100.0, value=70.0, step=0.1, format="%.1f")
        satis_fiyati = st.number_input("SatÄ±ÅŸ FiyatÄ± (50 KG) *", min_value=0.0, value=980.00, step=0.01, format="%.2f")
        belge = st.number_input("Belge Geliri (50 KG)", min_value=0.0, value=0.00, step=0.01, format="%.2f")

    with col2:
        st.markdown("#### ğŸ“Š YAN ÃœRÃœN ORANLARI (%)")
        col_oran1, col_oran2 = st.columns(2)
        with col_oran1:
            st.caption("Un OranÄ±")
            r_un2 = st.number_input("2. Un", min_value=0.0, value=7.0, step=0.1, format="%.1f", label_visibility="collapsed", key="r_un2")
            st.caption("Bongalite")
            r_bon = st.number_input("Bongalite %", min_value=0.0, value=1.5, step=0.1, format="%.1f", label_visibility="collapsed", key="r_bon")
        with col_oran2:
            st.caption("Kepek OranÄ±")
            r_kep = st.number_input("Kepek", min_value=0.0, value=9.0, step=0.1, format="%.1f", label_visibility="collapsed", key="r_kep")
            st.caption("Razmol OranÄ±")
            r_raz = st.number_input("Razmol", min_value=0.0, value=11.0, step=0.1, format="%.1f", label_visibility="collapsed", key="r_raz")
        
        st.markdown("#### ğŸ’° YAN ÃœRÃœN FÄ°YATLARI")
        col_fiyat1, col_fiyat2 = st.columns(2)
        with col_fiyat1:
            st.caption("Un Fiyat")
            p_un2 = st.number_input("2. Un TL", min_value=0.0, value=17.00, step=0.01, format="%.2f", label_visibility="collapsed", key="p_un2")
            st.caption("Bongalite Fiyat")
            p_bon = st.number_input("Bon. TL", min_value=0.0, value=11.60, step=0.01, format="%.2f", label_visibility="collapsed", key="p_bon")
        with col_fiyat2:
            st.caption("Kepek Fiyat")
            p_kep = st.number_input("Kepek TL", min_value=0.0, value=8.90, step=0.01, format="%.2f", label_visibility="collapsed", key="p_kep")
            st.caption("Razmol Fiyat")
            p_raz = st.number_input("Razmol TL", min_value=0.0, value=9.10, step=0.01, format="%.2f", label_visibility="collapsed", key="p_raz")
        
        st.markdown("#### ğŸŒ¾ EK GELÄ°RLER")
        col_ek1, col_ek2 = st.columns(2)
        with col_ek1:
            st.caption("SatÄ±lan KÄ±rÄ±k (Ton)")
            kirik_tonaj = st.number_input("KÄ±rÄ±k Ton", min_value=0.0, value=0.0, step=0.5, format="%.1f", label_visibility="collapsed", key="kirik_tonaj")
            st.caption("SatÄ±lan BaÅŸak (Ton)")
            basak_tonaj = st.number_input("BaÅŸak Ton", min_value=0.0, value=0.0, step=0.5, format="%.1f", label_visibility="collapsed", key="basak_tonaj")
        with col_ek2:
            st.caption("KÄ±rÄ±k Fiyat (TL/KG)")
            kirik_fiyat = st.number_input("KÄ±rÄ±k TL/KG", min_value=0.0, value=0.0, step=0.01, format="%.2f", label_visibility="collapsed", key="kirik_fiyat")
            st.caption("BaÅŸak Fiyat (TL/KG)")
            basak_fiyat = st.number_input("BaÅŸak TL/KG", min_value=0.0, value=0.0, step=0.01, format="%.2f", label_visibility="collapsed", key="basak_fiyat")

    with col3:
        st.markdown("#### ğŸ¢ AYLIK SABÄ°T GÄ°DERLER")
        g_personel = st.number_input("Personel MaaÅŸÄ±", min_value=0.0, value=1200000.0, step=1000.0, format="%.2f")
        g_bakim = st.number_input("BakÄ±m Maliyeti", min_value=0.0, value=100000.0, step=1000.0, format="%.2f")
        g_mutfak = st.number_input("Mutfak (Kantin)", min_value=0.0, value=50000.0, step=1000.0, format="%.2f")
        g_finans = st.number_input("Finans (Banka)", min_value=0.0, value=0.0, step=1000.0, format="%.2f")
        g_diger = st.number_input("DiÄŸer Giderler", min_value=0.0, value=0.0, step=1000.0, format="%.2f")
        
        st.markdown("#### âš¡ ELEKTRÄ°K")
        g_elektrik_birim = st.number_input("1 Ton BuÄŸday Elektrik (TL)", min_value=0.0, value=500.00, step=0.01)
        elektrik_aylik = g_elektrik_birim * aylik_kirilan
        st.caption(f"AylÄ±k Elektrik: {elektrik_aylik:,.0f} {currency}")
        
        st.markdown("#### ğŸ›’ Ã‡UVAL BAÅI GÄ°DERLER")
        col_cg1, col_cg2 = st.columns(2)
        with col_cg1:
            st.caption("Nakliye")
            g_nakliye = st.number_input("Nakliye Gider", min_value=0.0, value=20.00, step=0.5, label_visibility="collapsed", key="g_nakliye")
            st.caption("Pazarlama")
            g_pazarlama = st.number_input("Pazarlama Gider", min_value=0.0, value=20.50, step=0.5, label_visibility="collapsed", key="g_pazarlama")
        with col_cg2:
            st.caption("PP Ã‡uval")
            g_cuval = st.number_input("PP Ã‡uval Gider", min_value=0.0, value=15.00, step=0.5, label_visibility="collapsed", key="g_cuval")
            st.caption("Enzim/KatkÄ±")
            g_katki = st.number_input("KatkÄ± Gider", min_value=0.0, value=9.00, step=0.5, label_visibility="collapsed", key="g_katki")

    st.divider()
    if st.button("ğŸ§® HESAPLA VE KAYDET", type="primary", use_container_width=True):
        un_tonaj = aylik_kirilan * (randiman / 100)
        cuval_sayisi = (un_tonaj * 1000) / 50
        
        gelir_un = cuval_sayisi * satis_fiyati
        gelir_un2 = (aylik_kirilan * 1000) * (r_un2 / 100) * p_un2
        gelir_bon = (aylik_kirilan * 1000) * (r_bon / 100) * p_bon
        gelir_kep = (aylik_kirilan * 1000) * (r_kep / 100) * p_kep
        gelir_raz = (aylik_kirilan * 1000) * (r_raz / 100) * p_raz
        gelir_belge = belge * cuval_sayisi
        gelir_kirik = kirik_tonaj * kirik_fiyat
        gelir_basak = basak_tonaj * basak_fiyat
        toplam_gelir = gelir_un + gelir_un2 + gelir_bon + gelir_kep + gelir_raz + gelir_belge + gelir_kirik + gelir_basak
        
        gider_bugday = bugday_maliyet * aylik_kirilan * 1000
        gider_elektrik = elektrik_aylik
        gider_sabit = g_personel + g_bakim + g_mutfak + g_finans + g_diger
        gider_nakliye = g_nakliye * cuval_sayisi
        gider_pazarlama = g_pazarlama * cuval_sayisi
        gider_cuval = g_cuval * cuval_sayisi
        gider_katki = g_katki * cuval_sayisi
        toplam_gider = gider_bugday + gider_elektrik + gider_sabit + gider_nakliye + gider_pazarlama + gider_cuval + gider_katki
        
        net_kar = toplam_gelir - toplam_gider
        net_kar_cuval = net_kar / cuval_sayisi if cuval_sayisi > 0 else 0
        maliyet_fabrika = satis_fiyati - net_kar_cuval
        
        st.success("âœ… Hesaplama TamamlandÄ±!")
        m1, m2, m3 = st.columns(3)
        m1.metric("Net Kar (50kg)", f"{net_kar_cuval:.2f} TL")
        m2.metric("Fabrika Maliyet", f"{maliyet_fabrika:.2f} TL")
        m3.metric("Toplam Kar", f"{net_kar:,.0f} TL")
        
        data = {
            'ay': secilen_ay, 'yil': secilen_yil, 'un_cesidi': un_cesidi, 
            'bugday_pacal_maliyeti': bugday_maliyet, 'aylik_kirilan_bugday': aylik_kirilan,
            'un_randimani': randiman, 'un_satis_fiyati': satis_fiyati, 'belge_geliri': belge,
            'un2_orani': r_un2, 'un2_fiyati': p_un2, 'bongalite_orani': r_bon, 'bongalite_fiyati': p_bon,
            'kepek_orani': r_kep, 'kepek_fiyati': p_kep, 'razmol_orani': r_raz, 'razmol_fiyati': p_raz,
            'kirik_tonaj': kirik_tonaj, 'kirik_fiyat': kirik_fiyat, 'basak_tonaj': basak_tonaj, 'basak_fiyat': basak_fiyat,
            'ton_bugday_elektrik': g_elektrik_birim, 'elektrik_gideri': gider_elektrik,
            'personel_maasi': g_personel, 'bakim_maliyeti': g_bakim, 'mutfak_gideri': g_mutfak,
            'finans_gideri': g_finans, 'diger_giderler': g_diger,
            'nakliye': g_nakliye, 'satis_pazarlama': g_pazarlama, 'pp_cuval': g_cuval, 'katki_maliyeti': g_katki,
            'net_kar_50kg': net_kar_cuval, 'net_kar_kg': net_kar_cuval / 50,
            'fabrika_cikis_maliyet': maliyet_fabrika, 'net_kar_toplam': net_kar,
            'toplam_gelir': toplam_gelir, 'toplam_gider': toplam_gider
        }
        
        if save_un_maliyet(data):
            st.success(f"ğŸ’¾ KayÄ±t BaÅŸarÄ±lÄ±: {secilen_ay} {secilen_yil}")
            time.sleep(1.5)
            st.rerun()
        else:
            st.error("âŒ KayÄ±t BaÅŸarÄ±sÄ±z!")


def show_un_maliyet_gecmisi():
    """Maliyet GeÃ§miÅŸi - Dashboard"""
    st.header("ğŸ“Š Un Maliyet GeÃ§miÅŸi & Trendler")
    
    df = get_un_maliyet_gecmisi()
    
    if df.empty:
        st.info("ğŸ“­ HenÃ¼z maliyet hesaplamasÄ± kaydÄ± bulunmamaktadÄ±r.")
        st.info("ğŸ’¡ Ä°lk hesaplamayÄ± yapmak iÃ§in 'Un Maliyet Hesaplama' menÃ¼sÃ¼ne gidin.")
        return
    
    st.subheader("ğŸ“ˆ Ã–zet GÃ¶stergeler")
    son_kayit = df.iloc[0]
    ort_kar = df['net_kar_50kg'].mean() if 'net_kar_50kg' in df.columns else 0
    ort_maliyet = df['fabrika_cikis_maliyet'].mean() if 'fabrika_cikis_maliyet' in df.columns else 0
    
    kpi1, kpi2, kpi3, kpi4 = st.columns(4)
    with kpi1:
        st.metric("Son KayÄ±t: Net Kar (50kg)", f"{son_kayit.get('net_kar_50kg', 0):.2f} TL",
                  delta=f"{son_kayit.get('net_kar_50kg', 0) - ort_kar:.2f} TL" if ort_kar > 0 else None)
    with kpi2:
        st.metric("Son KayÄ±t: Fabrika Maliyet", f"{son_kayit.get('fabrika_cikis_maliyet', 0):.2f} TL",
                  delta=f"{son_kayit.get('fabrika_cikis_maliyet', 0) - ort_maliyet:.2f} TL" if ort_maliyet > 0 else None,
                  delta_color="inverse")
    with kpi3:
        st.metric("Son KayÄ±t: Toplam Kar", f"{son_kayit.get('net_kar_toplam', 0):,.0f} TL")
    with kpi4:
        st.metric("Toplam KayÄ±t SayÄ±sÄ±", f"{len(df)} Hesaplama")
    
    st.divider()
    st.subheader("ğŸ“‰ Trend Grafikleri")
    
    if 'tarih' in df.columns:
        df['tarih_str'] = df['tarih'].dt.strftime('%d/%m/%Y')
    
    tab1, tab2, tab3 = st.tabs(["ğŸ’° KarlÄ±lÄ±k Trendi", "ğŸ“Š Maliyet-SatÄ±ÅŸ", "ğŸ“ˆ AylÄ±k Performans"])
    
    with tab1:
        if 'net_kar_50kg' in df.columns and 'tarih_str' in df.columns:
            fig1 = px.line(df, x='tarih_str', y='net_kar_50kg', title="Ã‡uval BaÅŸÄ±na Net Kar",
                          labels={'tarih_str': 'Tarih', 'net_kar_50kg': 'Net Kar (TL)'}, markers=True)
            st.plotly_chart(fig1, use_container_width=True)
    
    with tab2:
        if 'fabrika_cikis_maliyet' in df.columns and 'un_satis_fiyati' in df.columns:
            fig2 = go.Figure()
            fig2.add_trace(go.Scatter(x=df['tarih_str'], y=df['fabrika_cikis_maliyet'], mode='lines+markers',
                                     name='Maliyet', line=dict(color='red')))
            fig2.add_trace(go.Scatter(x=df['tarih_str'], y=df['un_satis_fiyati'], mode='lines+markers',
                                     name='SatÄ±ÅŸ', line=dict(color='green')))
            fig2.update_layout(title="Maliyet vs SatÄ±ÅŸ", xaxis_title="Tarih", yaxis_title="Fiyat (TL)")
            st.plotly_chart(fig2, use_container_width=True)
    
    with tab3:
        if 'net_kar_toplam' in df.columns and 'tarih_str' in df.columns:
            fig3 = px.bar(df, x='tarih_str', y='net_kar_toplam', title="Toplam Kar",
                         color='net_kar_toplam', color_continuous_scale='RdYlGn')
            st.plotly_chart(fig3, use_container_width=True)
    
    st.divider()
    st.subheader("ğŸ“‹ DetaylÄ± KayÄ±tlar")
    
    display_cols = ['tarih_str', 'un_cesidi', 'net_kar_50kg', 'fabrika_cikis_maliyet',
                    'un_satis_fiyati', 'net_kar_toplam', 'aylik_kirilan_bugday', 'kullanici']
    display_cols = [c for c in display_cols if c in df.columns]
    
    df_display = df[display_cols].copy()
    df_display.columns = ['Tarih', 'Un Ã‡eÅŸidi', 'Net Kar (50kg)', 'Fabrika Maliyet',
                          'SatÄ±ÅŸ FiyatÄ±', 'Toplam Kar', 'KÄ±rÄ±lan (Ton)', 'KullanÄ±cÄ±'][:len(display_cols)]
    
    st.dataframe(df_display, use_container_width=True, hide_index=True, height=400)
    
    st.divider()
    if st.button("ğŸ“¥ Excel Ä°ndir", type="primary"):
        filename = f"un_maliyet_{datetime.now().strftime('%Y%m%d')}.xlsx"
        download_styled_excel(df, filename, "Maliyet GeÃ§miÅŸi")
        
    # SÄ°LME PANELÄ° (Sadece Admin GÃ¶rebilir)
    if st.session_state.get('user_role') == 'admin':
        st.divider()
        with st.expander("ğŸ—‘ï¸ KayÄ±t Silme Paneli (Test Verilerini Temizle)", expanded=False):
            st.warning("âš ï¸ Dikkat: Bu iÅŸlem geri alÄ±namaz!")
            
            # SeÃ§im Listesi (Tarih ve Un Ã‡eÅŸidi gÃ¶sterelim)
            secenekler = df.to_dict('records')
            
            def format_func_del(row):
                # GÃ¼venli gÃ¶sterim
                tarih = row.get('tarih_str', str(row.get('tarih')))
                un = row.get('un_cesidi', 'Bilinmiyor')
                kar = row.get('net_kar_50kg', 0)
                return f"{tarih} - {un} (Net Kar: {kar:.2f} TL)"

            silinecek_kayit = st.selectbox(
                "Silinecek KaydÄ± SeÃ§in:", 
                secenekler, 
                format_func=format_func_del,
                key="del_maliyet_select"
            )

            if silinecek_kayit:
                col_del_btn, col_del_info = st.columns([1, 4])
                with col_del_btn:
                    if st.button("ğŸ”¥ KaydÄ± Sil", type="primary", key="btn_del_confirm"):
                        # Orijinal 'tarih' verisini kullanarak sil
                        if delete_un_maliyet_record(silinecek_kayit['tarih']):
                            st.success("âœ… KayÄ±t baÅŸarÄ±yla silindi!")
                            time.sleep(1)
                            st.rerun() # Listeyi yenile
                        else:
                            st.error("âŒ Silme iÅŸlemi sÄ±rasÄ±nda hata oluÅŸtu.")
  
  
def show_flour_yonetimi():
    # 1. BaÅŸlÄ±k AlanÄ±
    st.markdown("""
    <div style='background-color: #FFF8E1; padding: 15px; border-radius: 10px; margin-bottom: 20px; border-left: 5px solid #FFB300;'>
        <h2 style='color: #E65100; margin:0;'>ğŸ Un Kalite Kontrol Merkezi</h2>
        <p style='color: #666; margin:0; font-size: 14px;'>Laboratuvar Analizleri, Standartlar ve AkÄ±llÄ± Dozajlama</p>
    </div>
    """, unsafe_allow_html=True)

    # 2. Yatay MenÃ¼ (Senin belirlediÄŸin profesyonel isimler)
    secim = st.radio(
        "ModÃ¼l SeÃ§iniz:",
        ["ğŸ“ Spek & Hedefler", "ğŸ§ª Analiz GiriÅŸi", "ğŸ“‚ Veri TabanÄ± & Rapor", "ğŸ’Š Enzim Dozaj Hesapla"],
        horizontal=True,
        label_visibility="collapsed"
    )
    
    st.markdown("---")

    # 3. YÃ¶nlendirmeler
    
    # --- A) SPEK & HEDEFLER ---
    if secim == "ğŸ“ Spek & Hedefler":
        # Yetki KontrolÃ¼
        user_role = st.session_state.get('user_role', 'viewer')
        
        if user_role == 'admin':
            with st.container(border=True):
                st.success("ğŸ”“ **YÃ¶netici Modu:** Kalite hedeflerini dÃ¼zenleyebilirsiniz.")
                show_spec_yonetimi()
        else:
            # Admin deÄŸilse uyarÄ± ver
            with st.container(border=True):
                st.warning("ğŸ”’ **Salt Okunur:** Kalite hedeflerini sadece YÃ¶neticiler deÄŸiÅŸtirebilir. Åu an sadece gÃ¶rÃ¼ntÃ¼lÃ¼yorsunuz.")
                show_spec_yonetimi()

    # --- B) ANALÄ°Z GÄ°RÄ°ÅÄ° ---
    elif secim == "ğŸ§ª Analiz GiriÅŸi":
        with st.container(border=True):
            show_un_analiz_kaydi()

    # --- C) VERÄ° TABANI & RAPOR ---
    elif secim == "ğŸ“‚ Veri TabanÄ± & Rapor":
        with st.container(border=True):
            show_un_analiz_kayitlari()

    # --- D) ENZÄ°M DOZAJ ---
    elif secim == "ğŸ’Š Enzim Dozaj Hesapla":
        with st.container(border=True):
            try:
                # Ä°sim Ã§akÄ±ÅŸmasÄ±nÄ± Ã¶nlemek iÃ§in 'as calc_module' dedik
                import app.modules.calculations as calc_module
                calc_module.show_enzim_dozajlama()
            except ImportError:
                st.error("âš ï¸ Enzim modÃ¼lÃ¼ (calculations.py) bulunamadÄ±.")
            except Exception as e:
                st.error(f"âš ï¸ ModÃ¼l yÃ¼klenirken hata oluÅŸtu: {e}")






































